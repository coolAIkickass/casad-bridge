"""
Parse CASAD E2E design input Excel (BBS sheet).
Returns a structured dict with geometry + BBS per component.
"""
import io
import logging
import openpyxl

log = logging.getLogger(__name__)


UNIT_WEIGHTS = {8: 0.395, 10: 0.617, 12: 0.888, 16: 1.578, 20: 2.467, 25: 3.857, 32: 6.313}

BAR_MARK_COL   = 0   # A
REMARKS_COL    = 4   # E
LENGTH_COL     = 7   # H
DIA_COL        = 8   # I
SPACING_COL    = 9   # J
COUNT_COL      = 10  # K
TOTAL_LEN_COL  = 11  # L
UNIT_WT_COL    = 12  # M
TOTAL_WT_COL   = 13  # N

KNOWN_BAR_MARKS = set(list('abcdefghijklmnopqrstuvwxyz') + [
    'a1','b1','c1','d1','e1','f1','g1','h1','i1','j1','k1',
    'x1','x2','y1','y2','z1',
])


def _is_bar_row(row_vals):
    mark = row_vals[BAR_MARK_COL]
    if not isinstance(mark, str):
        return False
    return mark.strip().lower() in KNOWN_BAR_MARKS


def _safe_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _parse_spacing(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if float(v) > 0 else None
    s = str(v).strip()
    if s in ('-', '', 'None'):
        return None
    # e.g. '10x20' → not a spacing, it's a count notation for stirrups
    if 'x' in s.lower():
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_bar_row(row_vals, shape_cols=None):
    mark = str(row_vals[BAR_MARK_COL]).strip()
    shape_dims = None
    if shape_cols:
        dims = []
        for ci in shape_cols:
            if ci < len(row_vals):
                v = _safe_float(row_vals[ci])
                # Only keep plausible bar dimension values (1mm–50000mm)
                if v is not None and v > 0:
                    dims.append(v)
        if dims:
            shape_dims = dims
    return {
        'bar_mark':    mark,
        'length_m':    _safe_float(row_vals[LENGTH_COL]),
        'dia_mm':      int(row_vals[DIA_COL]) if row_vals[DIA_COL] else None,
        'spacing_mm':  _parse_spacing(row_vals[SPACING_COL]),
        'count':       int(row_vals[COUNT_COL]) if row_vals[COUNT_COL] is not None else None,
        'total_len_m': _safe_float(row_vals[TOTAL_LEN_COL]),
        'unit_wt':     _safe_float(row_vals[UNIT_WT_COL]),
        'total_wt_kg': _safe_float(row_vals[TOTAL_WT_COL]),
        'remarks':     str(row_vals[REMARKS_COL]).strip() if row_vals[REMARKS_COL] else '',
        'raw_spacing': str(row_vals[SPACING_COL]) if row_vals[SPACING_COL] is not None else None,
        'shape_dims':  shape_dims,
    }


def parse_e2e_excel(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.worksheets[0]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    # Open a second time without data_only to read raw formulas for pile count detection.
    wb_formula = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    ws_formula = wb_formula.worksheets[0]
    formula_rows = [list(r) for r in ws_formula.iter_rows(values_only=True)]

    geometry = _parse_geometry(rows)
    pilecap_bbs, pile_bbs, pier_bbs = _parse_bbs_sections(rows, formula_rows, geometry)

    return {
        'source': 'e2e_excel',
        'geometry': geometry,
        'pilecap_bbs': pilecap_bbs,
        'pile_bbs':    pile_bbs,
        'pier_bbs':    pier_bbs,
    }


def _parse_geometry(rows):
    geo = {}
    for row in rows:
        for ci, cell in enumerate(row):
            if not isinstance(cell, str):
                continue
            c = cell.strip()
            if 'Length of Pilecap (along traffic)' in c:
                geo['pilecap_length_along'] = _safe_float(row[ci + 1])
                if row[ci + 2] == 'a1=':
                    geo['pilecap_a1'] = _safe_float(row[ci + 3])
            elif 'Length of Pilecap (across traffic)' in c:
                geo['pilecap_length_across'] = _safe_float(row[ci + 1])
                if row[ci + 2] == 'b1=':
                    geo['pilecap_b1'] = _safe_float(row[ci + 3])
            elif 'Depth of Pilecap' in c:
                geo['pilecap_depth'] = _safe_float(row[ci + 1])
            elif 'Foundation Cover' in c or 'Foundation cover' in c:
                geo['pilecap_cover'] = _safe_float(row[ci + 1])
            elif 'Pile Dia=' in c:
                geo['pile_dia'] = _safe_float(row[ci + 1])
            elif 'No. of Piles=' in c:
                geo['pile_count'] = int(row[ci + 1]) if row[ci + 1] else None
            elif 'Pile length=' in c:
                geo['pile_length'] = _safe_float(row[ci + 1])
                if row[ci + 2] == 'Pile Fix. length=':
                    geo['pile_fixity'] = _safe_float(row[ci + 3])
            elif 'Pier Shape=' in c and 'Rectangle' in c:
                geo['pier_shape'] = 'Rectangle'
            elif 'Pier Shape=' in c and 'Circular' in c:
                geo['pier_shape'] = 'Circular'
            elif 'Ht. of Pier=' in c:
                geo['pier_height'] = _safe_float(row[ci + 1])
            elif 'Pier Dimsion' in c or 'Pier Dimension' in c:
                geo['pier_length'] = _safe_float(row[ci + 1])
                geo['pier_width'] = _safe_float(row[ci + 2]) if len(row) > ci + 2 else None
            elif 'Pier Cover' in c:
                geo['pier_cover'] = _safe_float(row[ci + 1])
    return geo


def _num_piles_cell_ref(rows) -> str | None:
    """Return the Excel cell address (e.g. 'D8') of the 'No. of Piles=' value cell."""
    for ri, row in enumerate(rows):
        for ci, cell in enumerate(row):
            if isinstance(cell, str) and 'No. of Piles=' in cell:
                # The value is in the next column; convert to Excel address (1-indexed col, 1-indexed row)
                col_letter = chr(ord('A') + ci + 1)
                return f'{col_letter}{ri + 1}'
    return None


def _formula_references_cell(formula_value, cell_ref: str) -> bool:
    """Return True if formula_value (raw cell value from data_only=False) references cell_ref."""
    if not isinstance(formula_value, str) or not formula_value.startswith('='):
        return False
    return cell_ref.upper() in formula_value.upper()


def _find_shape_cols(rows) -> list:
    """Scan rows for the 'Shape of bar' column header and return the column indices
    spanning that header up to (but not including) REMARKS_COL.
    Searches for 'SHAPE OF' to avoid matching geometry cells like 'Pier Shape=Rectangle'."""
    for row in rows:
        for ci, cell in enumerate(row):
            if isinstance(cell, str) and 'SHAPE OF' in cell.upper():
                # Header found (e.g. "Shape of bar") at column ci.
                # Shape dim cells occupy ci through REMARKS_COL-1.
                cols = [c for c in range(ci, REMARKS_COL) if c != BAR_MARK_COL]
                log.info('Shape column header found at col %d; scanning cols %s', ci, cols)
                return cols
    log.info('No "Shape of bar" header found — shape dimension check will be skipped')
    return []


def _parse_bbs_sections(rows, formula_rows=None, geometry=None):
    pilecap_bbs = {}
    pile_bbs    = {}
    # Collect both pier types separately, then pick the right one
    pier_circular    = {}
    pier_rectangular = {}

    current      = None
    active_pier  = None  # 'Circular' or 'Rectangle'
    stirrup_spacing = None
    pier_shape_from_geo = None

    num_piles = (geometry or {}).get('pile_count') or 1
    # Find the cell address of num_piles so we can detect formula references to it
    pile_count_ref = _num_piles_cell_ref(rows) if rows else None

    # Locate shape-of-bar dimension columns dynamically by header label
    shape_cols = _find_shape_cols(rows)

    # First pass: detect pier shape from geometry section
    for row in rows:
        for ci, cell in enumerate(row):
            if isinstance(cell, str) and 'Pier Shape=' in cell:
                if 'Rectangle' in cell:
                    pier_shape_from_geo = 'Rectangle'
                elif 'Circular' in cell:
                    pier_shape_from_geo = 'Circular'

    for ri, row in enumerate(rows):
        mark_cell = row[BAR_MARK_COL]
        if isinstance(mark_cell, str):
            label = mark_cell.strip().upper()
            if label == 'PILECAP':
                current = 'pilecap'
                continue
            elif label.startswith('PILE ('):
                current = 'pile'
                continue
            elif 'PIER (CIRCULAR)' in label:
                current = 'pier'
                active_pier = 'Circular'
                continue
            elif 'PIER (RECTANGULAR' in label or 'PIER (CAPSULE' in label:
                current = 'pier'
                active_pier = 'Rectangle'
                continue
            elif label.startswith('NOTE :'):
                continue

        if current and row[2] is not None and isinstance(row[2], str):
            if 'Spacing b/w long' in str(row[2]):
                stirrup_spacing = _safe_float(row[7])

        if current and _is_bar_row(row):
            bar = _parse_bar_row(row, shape_cols=shape_cols)
            # Skip inactive bars (dia = 0 or count = 0)
            if (not bar['dia_mm']) or bar['dia_mm'] == 0:
                continue
            if bar.get('count') == 0:
                continue
            # Attach stirrup longitudinal spacing for 'e'
            if bar['spacing_mm'] is None and stirrup_spacing and bar['bar_mark'] == 'e':
                bar['spacing_mm'] = stirrup_spacing

            # For PILE bars: normalise count to total (across all piles).
            # The Excel section is "per pile". Check whether the count cell formula
            # already references the num_piles cell. If it does → already total.
            # If it doesn't (or is a plain number) → per-pile → multiply by num_piles.
            if current == 'pile' and bar.get('count') is not None:
                count_formula = (
                    (formula_rows[ri][COUNT_COL] if formula_rows else None)
                    if ri < len(formula_rows or []) else None
                )
                already_total = (
                    pile_count_ref
                    and _formula_references_cell(count_formula, pile_count_ref)
                )
                if not already_total and num_piles > 1:
                    bar['count'] = bar['count'] * num_piles
                    if bar.get('total_len_m') is not None:
                        bar['total_len_m'] = bar['total_len_m'] * num_piles
                    if bar.get('total_wt_kg') is not None:
                        bar['total_wt_kg'] = bar['total_wt_kg'] * num_piles
            elif current == 'pile' and bar.get('count') is None:
                log.warning(
                    'Pile bar %s: count cell is None — formula cache may be stale. '
                    'Re-save the Excel in Microsoft Excel to refresh formula results.',
                    bar['bar_mark']
                )

            bm = bar['bar_mark']

            def _add(d, bm, bar):
                if bm in d:
                    ex = d[bm]
                    if isinstance(ex, list):
                        ex.append(bar)
                    else:
                        d[bm] = [ex, bar]
                else:
                    d[bm] = bar

            if current == 'pilecap':
                _add(pilecap_bbs, bm, bar)
            elif current == 'pile':
                _add(pile_bbs, bm, bar)
            elif current == 'pier':
                if active_pier == 'Circular':
                    _add(pier_circular, bm, bar)
                else:
                    _add(pier_rectangular, bm, bar)

    # Select the correct pier BBS based on geometry
    if pier_shape_from_geo == 'Circular':
        pier_bbs = pier_circular
    elif pier_shape_from_geo == 'Rectangle':
        pier_bbs = pier_rectangular
    else:
        # Default: prefer rectangular if both exist, else whichever has data
        pier_bbs = pier_rectangular if pier_rectangular else pier_circular

    return pilecap_bbs, pile_bbs, pier_bbs
