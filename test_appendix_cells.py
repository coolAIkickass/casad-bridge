"""
test_appendix_cells.py — Senior-tester regression suite for Appendix-A and Appendix-B
cell values in both R&B and AMC modules.

Each test case is tagged with:
  Module  : RB | AMC
  Sheet   : A | B
  Cell    : e.g. C15
  BUG-ID  : sequential identifier
  Expected: exact value the template row label demands
  Failure mode: what the old code wrote instead

Run:
    cd /Users/bansal.umang89/Desktop/CASAD/casad-bridge
    python -m pytest test_appendix_cells.py -v
"""
import json
import os
import sys
import unittest
import openpyxl
from openpyxl.utils import coordinate_to_tuple

# ── locate fixture ────────────────────────────────────────────────────────────
FIXTURE = os.path.join(os.path.dirname(__file__), 'test_fixture_khokhara.json')


def _load():
    with open(FIXTURE, encoding='utf-8') as f:
        return json.load(f)


def _wb(*sheet_names):
    """Minimal workbook with specified sheet names for mock-fill tests."""
    wb = openpyxl.Workbook()
    wb.active.title = sheet_names[0]
    for name in sheet_names[1:]:
        wb.create_sheet(name)
    return wb


def _get(ws, addr):
    r, c = coordinate_to_tuple(addr)
    return ws.cell(row=r, column=c).value


# ══════════════════════════════════════════════════════════════════════════════
#  R&B MODULE — Appendix-A
# ══════════════════════════════════════════════════════════════════════════════
class TestRB_AppendixA(unittest.TestCase):
    """
    Ground truth: casad_excel_template.xlsx, sheet 'Appendix-A'
    Template row labels verified 2026-05-18.
    """
    @classmethod
    def setUpClass(cls):
        from report_gen_excel import _fill_appendix_a
        cls.d  = _load()
        cls.wb = _wb('Appendix-A')
        _fill_appendix_a(cls.wb, cls.d)
        cls.ws = cls.wb['Appendix-A']

    # ── identity section ──────────────────────────────────────────────────────

    def test_TC01_C4_bridge_name(self):
        """BUG: none. Baseline check — name must match fixture."""
        self.assertEqual(_get(self.ws, 'C4'), 'KHOKHARA ROB')

    def test_TC02_C6_river_name(self):
        """BUG: none. Must say 'Not Applicable...'."""
        self.assertEqual(_get(self.ws, 'C6'),
                         'Not Applicable as it is Railway Over Bridge')

    def test_TC03_C9_gps_format(self):
        """BUG: none. Degree symbol must be present."""
        val = str(_get(self.ws, 'C9'))
        self.assertIn('23.007695°', val)
        self.assertIn('72.607542°', val)

    def test_TC04_C11_division(self):
        self.assertEqual(_get(self.ws, 'C11'), 'AMC')

    def test_TC05_C12_circle(self):
        self.assertEqual(_get(self.ws, 'C12'), 'AMC')

    # ── Details of Spans section ──────────────────────────────────────────────

    def test_TC06_C15_no_of_spans_AND_span_length(self):
        """
        C15 must use 4 fixed bold sub-titles:
          Number of Span / Length / C/C of Piers / Width of Piers
        Each sub-title is followed by the matching field value (only shown if data given).
        """
        val = str(_get(self.ws, 'C15') or '')
        # Fixed sub-title labels
        self.assertIn('Number of Span:', val, "Missing 'Number of Span:' sub-title")
        self.assertIn('Length:', val,         "Missing 'Length:' sub-title")
        self.assertIn('C/C of Piers:', val,   "Missing 'C/C of Piers:' sub-title")
        self.assertIn('Width of Piers:', val, "Missing 'Width of Piers:' sub-title")
        # Data content
        self.assertIn('Anupam Road Side: 10 Nos.', val, "no_of_spans missing from C15")
        self.assertIn('339.53', val,  "total_length missing from C15 Length sub-title")
        self.assertIn('25 m (Anupam Cinema Side)', val, "cc_of_piers missing from C15")
        self.assertIn('1.25 m (Anupam Cinema Side)', val, "width_of_piers missing from C15")

    def test_TC07_C16_total_length_math_preserved(self):
        """BUG: none. Math expression must survive verbatim."""
        val = str(_get(self.ws, 'C16') or '')
        self.assertIn('92 + 6 × 25', val)
        self.assertIn('339.53', val)
        self.assertIn('330 m', val)

    def test_TC08_C17_angle_of_crossing_is_Q(self):
        """
        BUG: template had stale 'Skew'. After pre-wipe+write, must be 'Q'.
        AI was converting 'Q' to 'Skew' (fixed in ai_parse.py prompt).
        """
        self.assertEqual(_get(self.ws, 'C17'), 'Q')

    def test_TC09_C18_bridge_level_type_ROB(self):
        """ROB High Level must appear — user stated type explicitly."""
        val = str(_get(self.ws, 'C18') or '')
        self.assertIn('ROB', val)
        self.assertIn('High Level', val)

    # ── Design and Structural Data section ───────────────────────────────────

    def test_TC10_C43_structural_type_all_three(self):
        """C43 = row-(b) 'Type of Bridge'. Must list PSC, RCC, and Steel Truss."""
        val = str(_get(self.ws, 'C43') or '')
        self.assertIn('PSC Girder', val)
        self.assertIn('RCC Solid Slab', val)
        self.assertIn('Steel Truss', val)

    def test_TC11_C44_span_arrangement_contains_railway(self):
        """C44 = row-(c) 'Span arrangement'."""
        val = str(_get(self.ws, 'C44') or '')
        self.assertIn('4 Nos.', val)
        self.assertIn('25 m', val)

    def test_TC12_C45_carriage_width_not_fabricated(self):
        """
        BUG: user never provided carriage width — must NOT be fabricated.
        Acceptable values: '-', None, or ''.
        """
        val = _get(self.ws, 'C45')
        self.assertIn(str(val) if val is not None else '-', ['-', ''],
                      f"Carriage width fabricated: {val!r}")

    def test_TC13_C46_deck_level(self):
        """C46 = row-(e) 'Deck level'. User gave 107.575 m."""
        val = str(_get(self.ws, 'C46') or '')
        self.assertIn('107.575', val)

    def test_TC14_C50_foundation_pile(self):
        val = str(_get(self.ws, 'C50') or '')
        self.assertIn('Pile', val)

    def test_TC15_C52_substructure_rcc_straight(self):
        """C52 = row-(i) 'Masonry, Mass Concrete, RCC'."""
        self.assertEqual(_get(self.ws, 'C52'), 'RCC Straight')

    def test_TC16_C53_pier_length_both_sides(self):
        """C53 = row-(ii) 'Straight length of pier'. Both values required."""
        val = str(_get(self.ws, 'C53') or '')
        self.assertIn('2.5', val)   # Anupam Cinema
        self.assertIn('3.0', val)   # Railway Portion

    def test_TC17_C67_expansion_joint(self):
        self.assertEqual(_get(self.ws, 'C67'), 'Strip Seal')

    # ── Other Data section ────────────────────────────────────────────────────

    def test_TC18_C81_date_completion_is_string_ddmmyyyy(self):
        """
        BUG-02 (R&B): _parse_survey_date() returned a datetime object.
        openpyxl writes it as '2020-08-09 00:00:00' or an Excel serial.
        Expected: string '09/08/2020'.
        """
        val = _get(self.ws, 'C81')
        self.assertIsInstance(val, str,
            f"C81 must be a str, got {type(val).__name__}: {val!r}")
        self.assertEqual(val, '09/08/2020',
            f"Expected '09/08/2020', got {val!r}")

    def test_TC19_C82_surface_utilities(self):
        self.assertEqual(_get(self.ws, 'C82'), 'Electric Lines etc.')

    def test_TC20_C92_performance(self):
        self.assertEqual(_get(self.ws, 'C92'), 'Good')

    def test_TC21_C93_date_survey_is_string_ddmmyyyy(self):
        """
        BUG-03 (R&B): same datetime-object issue as C81.
        Expected: string '13/04/2026'.
        """
        val = _get(self.ws, 'C93')
        self.assertIsInstance(val, str,
            f"C93 must be a str, got {type(val).__name__}: {val!r}")
        self.assertEqual(val, '13/04/2026',
            f"Expected '13/04/2026', got {val!r}")

    def test_TC22_template_pollution_wiped(self):
        """BUG: stale template values must not survive the pre-wipe."""
        # C20 is inside the hydraulic-parameters section — never written.
        # Must be None after wipe.
        self.assertIsNone(_get(self.ws, 'C20'),
            "C20 not cleared — stale template data leaking")


# ══════════════════════════════════════════════════════════════════════════════
#  R&B MODULE — Appendix-B
# ══════════════════════════════════════════════════════════════════════════════
class TestRB_AppendixB(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        from report_gen_excel import _fill_appendix_b
        cls.d  = _load()
        cls.wb = _wb('Appendix-B')
        _fill_appendix_b(cls.wb, cls.d)
        cls.ws = cls.wb['Appendix-B']

    def test_TC23_C12_date_survey_is_string(self):
        """
        BUG-04 (R&B Appendix-B): _parse_survey_date() returned datetime object.
        Expected: '13/04/2026'.
        """
        val = _get(self.ws, 'C12')
        self.assertIsInstance(val, str,
            f"Appendix-B C12 must be str, got {type(val).__name__}: {val!r}")
        self.assertEqual(val, '13/04/2026')

    def test_TC24_C10_div_equals_circle_no_slash(self):
        """When division == circle, write single value, not 'AMC / AMC'."""
        self.assertEqual(_get(self.ws, 'C10'), 'AMC')

    def test_TC25_C11_type_of_bridge(self):
        val = str(_get(self.ws, 'C11') or '')
        self.assertIn('ROB', val)


# ══════════════════════════════════════════════════════════════════════════════
#  AMC MODULE — Appendix-A
# ══════════════════════════════════════════════════════════════════════════════
class TestAMC_AppendixA(unittest.TestCase):
    """
    Ground truth: casad_amc_template.xlsx, sheet 'Appendix-A'
    BOTH templates have identical row structure (confirmed 2026-05-18).
    """
    @classmethod
    def setUpClass(cls):
        from report_gen_excel_amc import _fill_appendix_a
        cls.d  = _load()
        cls.wb = _wb('Appendix-A')
        _fill_appendix_a(cls.wb, cls.d)
        cls.ws = cls.wb['Appendix-A']

    def test_TC26_C14_is_blank_section_header(self):
        """
        BUG-05 (AMC): old code wrote no_of_spans to C14.
        C14 is 'Details of Spans:' section header — must stay blank.
        """
        val = _get(self.ws, 'C14')
        self.assertIsNone(val,
            f"C14 is the 'Details of Spans' section header and must be None, "
            f"got {val!r}")

    def test_TC27_C15_no_of_spans_AND_span_length(self):
        """
        AMC C15 must use 4 fixed bold sub-titles:
          Number of Span / Length / C/C of Piers / Width of Piers
        Each sub-title is followed by the matching field value (only shown if data given).
        """
        val = str(_get(self.ws, 'C15') or '')
        # Fixed sub-title labels
        self.assertIn('Number of Span:', val, "Missing 'Number of Span:' sub-title in AMC C15")
        self.assertIn('Length:', val,         "Missing 'Length:' sub-title in AMC C15")
        self.assertIn('C/C of Piers:', val,   "Missing 'C/C of Piers:' sub-title in AMC C15")
        self.assertIn('Width of Piers:', val, "Missing 'Width of Piers:' sub-title in AMC C15")
        # Data content
        self.assertIn('Anupam Road Side: 10 Nos.', val, "no_of_spans not at C15")
        self.assertIn('339.53', val,  "total_length missing from AMC C15 Length sub-title")
        self.assertIn('25 m (Anupam Cinema Side)', val, "cc_of_piers missing from AMC C15")
        self.assertIn('1.25 m (Anupam Cinema Side)', val, "width_of_piers missing from AMC C15")

    def test_TC28_C16_total_length(self):
        """
        BUG-07 (AMC): old code wrote total_length to C15 (number-of-spans row).
        Must be at C16.
        """
        val = str(_get(self.ws, 'C16') or '')
        self.assertIn('339.53', val, "total_length not at C16")
        self.assertIn('330 m', val)

    def test_TC29_C17_angle_of_crossing(self):
        """BUG-08 (AMC): C17 was never written. Must contain 'Q'."""
        self.assertEqual(_get(self.ws, 'C17'), 'Q',
                         "C17 angle_of_crossing missing in AMC")

    def test_TC30_C18_bridge_level_type(self):
        """BUG-09 (AMC): C18 was never written."""
        val = str(_get(self.ws, 'C18') or '')
        self.assertIn('ROB', val, "C18 bridge_level_type missing in AMC")

    def test_TC31_C43_structural_type_not_C44(self):
        """
        BUG-10 (AMC): old code wrote superstructure_type to C44 (span-arrangement row).
        Must be at C43 = row-(b) 'Type of Bridge'.
        """
        val = str(_get(self.ws, 'C43') or '')
        self.assertIn('PSC Girder', val,
                      "superstructure_type must be at C43, not C44")

    def test_TC32_C44_span_arrangement_not_C45(self):
        """
        BUG-11 (AMC): old code wrote span_arrangement to C45 (carriage-width row).
        Must be at C44 = row-(c) 'Span arrangement'.
        """
        val = str(_get(self.ws, 'C44') or '')
        self.assertTrue(
            '4 Nos.' in val or '25 m' in val,
            f"span_arrangement must be at C44, got {val!r}")

    def test_TC33_C45_carriage_width_not_fabricated(self):
        """
        BUG-12 (AMC): old code wrote span_arrangement here.
        C45 = row-(d) 'Carriage width and footpath width'.
        User never provided carriage width — must be '-' or blank.
        """
        val = _get(self.ws, 'C45')
        # Must NOT contain span-arrangement content
        self.assertNotIn('Nos.', str(val or ''),
                         f"span_arrangement leaked into carriage-width cell C45: {val!r}")
        self.assertIn(str(val) if val is not None else '-', ['-', ''],
                      f"Carriage width must not be fabricated: {val!r}")

    def test_TC34_C46_deck_level(self):
        """BUG-13 (AMC): C46 was never written."""
        val = str(_get(self.ws, 'C46') or '')
        self.assertIn('107.575', val, "deck_level not written to C46 in AMC")

    def test_TC35_C52_substructure_type(self):
        """BUG-14 (AMC): C52 never written."""
        self.assertEqual(_get(self.ws, 'C52'), 'RCC Straight',
                         "substructure_type missing from C52 in AMC")

    def test_TC36_C53_pier_length(self):
        """BUG-15 (AMC): C53 never written."""
        val = str(_get(self.ws, 'C53') or '')
        self.assertIn('2.5', val, "pier_length missing from C53 in AMC")

    def test_TC37_C59_superstructure_not_C61(self):
        """
        BUG-16 (AMC): old code wrote superstructure_type to C61
        (row-(j)(iii) 'Details of articulation').
        Must be at C59 = row-(j)(i) 'Type of superstructure'.
        """
        val = str(_get(self.ws, 'C59') or '')
        self.assertIn('PSC Girder', val,
                      "superstructure_type must be at C59, not C61")
        # C61 must NOT contain PSC (it's articulation details)
        c61 = str(_get(self.ws, 'C61') or '')
        self.assertNotIn('PSC Girder', c61,
                         f"superstructure_type leaked into C61: {c61!r}")

    def test_TC38_C60_prestressing_details(self):
        """BUG-17 (AMC): C60 never written."""
        val = _get(self.ws, 'C60')
        self.assertIsNotNone(val, "C60 prestressing_details missing in AMC")
        self.assertIn('Design', str(val))

    def test_TC39_C81_date_completion_string(self):
        """BUG-18 (AMC): C81 never written at all."""
        val = _get(self.ws, 'C81')
        self.assertIsNotNone(val, "C81 date_of_completion missing in AMC")
        self.assertEqual(str(val), '09/08/2020',
                         f"C81 expected '09/08/2020', got {val!r}")

    def test_TC40_C82_surface_utilities(self):
        """BUG-19 (AMC): C82 never written."""
        self.assertEqual(_get(self.ws, 'C82'), 'Electric Lines etc.',
                         "surface_utilities missing from C82 in AMC")

    def test_TC41_C92_performance(self):
        """BUG-20 (AMC): C92 never written."""
        self.assertEqual(_get(self.ws, 'C92'), 'Good',
                         "performance missing from C92 in AMC")

    def test_TC42_C93_date_survey_string(self):
        """BUG-21 (AMC): C93 never written."""
        val = _get(self.ws, 'C93')
        self.assertIsNotNone(val, "C93 date_of_survey missing in AMC")
        self.assertEqual(str(val), '13/04/2026',
                         f"C93 expected '13/04/2026', got {val!r}")


# ══════════════════════════════════════════════════════════════════════════════
#  AMC MODULE — Appendix-B
# ══════════════════════════════════════════════════════════════════════════════
class TestAMC_AppendixB(unittest.TestCase):
    """
    Both templates have IDENTICAL Appendix-B row structure.
    AMC _fill_appendix_b had three families of wrong row numbers.
    """
    @classmethod
    def setUpClass(cls):
        from report_gen_excel_amc import _fill_appendix_b
        cls.d  = _load()
        cls.wb = _wb('Appendix-B')
        _fill_appendix_b(cls.wb, cls.d)
        cls.ws = cls.wb['Appendix-B']

    def test_TC43_C14_approach_settlement_correct_row(self):
        """
        BUG-22 (AMC Appendix-B): old code wrote approach_settlement to C16.
        Template row 4.1 'condition of pavement' is at C14.
        """
        val = _get(self.ws, 'C14')
        self.assertIsNotNone(val,
            "approach_settlement must be at C14, was at C16")
        # Confirm C16 does NOT hold the approach-settlement text
        # (it should hold approach_erosion)

    def test_TC44_C16_approach_erosion_not_settlement(self):
        """
        BUG-22 continued: C16 must be approach_erosion (row 4.3), NOT settlement.
        """
        val = str(_get(self.ws, 'C16') or '')
        # approach_erosion = 'Not Visible'
        self.assertEqual(val, 'Not Visible',
            f"C16 must be approach_erosion 'Not Visible', got {val!r}")

    def test_TC45_C19_approach_other(self):
        """
        Row 4.6 'Any other specific observations' is at C19.
        Old code wrote this to C21.
        """
        val = str(_get(self.ws, 'C19') or '')
        self.assertIn('accessible', val.lower(),
            f"approach_other must be at C19, got {val!r}")

    def test_TC46_C52_sub_cracks_correct_row(self):
        """
        BUG-23 (AMC Appendix-B): old code wrote sub_cracks to C43
        (row 8.1 'drainage backfill' — wrong row).
        Row 9.1.4 'cracks in supporting member' is C52.
        """
        val = str(_get(self.ws, 'C52') or '')
        self.assertIn('crack', val.lower(),
            f"sub_cracks must be at C52, got {val!r}")

    def test_TC47_C43_not_sub_cracks(self):
        """C43 = row 8.1 drainage backfill — must NOT contain pedestal crack info."""
        val = str(_get(self.ws, 'C43') or '')
        self.assertNotIn('pedestal', val.lower(),
            f"sub_cracks leaked into drainage-backfill row C43: {val!r}")

    def test_TC48_C59_superstructure_refer_table(self):
        """
        BUG-24 (AMC Appendix-B): old code wrote ss_cracks to C55
        (row 9.2 'Elastomeric bearings' section header — wrong).
        Row 10 'Superstructure' is C59.
        """
        val = str(_get(self.ws, 'C59') or '')
        # Should contain a reference to the defect tables
        self.assertIn('Table', val,
            f"C59 superstructure row must reference tables, got {val!r}")

    def test_TC49_C55_not_ss_cracks(self):
        """C55 = row 9.2 Elastomeric bearings — must NOT contain SS crack data."""
        val = str(_get(self.ws, 'C55') or '')
        # ss_cracks content should NOT be here
        crack_text = self.d.get('ss_cracks', '')
        if crack_text and crack_text != '-':
            self.assertNotEqual(val, crack_text,
                f"ss_cracks leaked into elastomeric-bearings row C55: {val!r}")


if __name__ == '__main__':
    unittest.main(verbosity=2)
