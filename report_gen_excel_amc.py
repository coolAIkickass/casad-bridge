# report_gen_excel_amc.py — Fill CASAD AMC Excel template from report JSON
#
# AMC template structure:
#   - TITLE PAGE, DISCLAIMR, Appendix-A, Appendix-B, Appendix-B 21-22,
#     Appendix - Photos, Table 1, Table 2, Table 3, Table 4
#   - Defect sheets: Table 1 (SUB side 1), Table 2 (SUB side 2),
#                    Table 3 (SUPER side 1), Table 4 (SUPER side 2)
#   - All defect tables: row 3 = pier/span IDs starting col 5 (E),
#                        rows 4-14 = defect observations (a-k)
#   - Remarks column detected dynamically from row 3
#   - Single photo sheet: "Appendix - Photos"

import os, re, io
from datetime import date, datetime
import openpyxl
from openpyxl.drawing.image import Image as XLImage

AMC_TEMPLATE_PATH = os.getenv('AMC_TEMPLATE_PATH', 'casad_amc_template.xlsx')
OUTPUT_DIR = os.getenv('OUTPUT_DIR', 'media')

# Cache raw template bytes at import time — avoids repeated disk reads under load.
# Each build_excel_amc() call loads from memory via io.BytesIO.
try:
    with open(AMC_TEMPLATE_PATH, 'rb') as _f:
        _AMC_TEMPLATE_BYTES = _f.read()
except FileNotFoundError:
    _AMC_TEMPLATE_BYTES = None

DEFECT_ROW = {
    'cracks': 4, 'leaching': 5, 'honeycombing': 6, 'exposed_rebar': 7,
    'leakage': 8, 'spalling': 9, 'rust_marks': 10, 'shuttering': 11,
    'delamination': 12, 'vegetation': 13, 'any_other': 14
}


def _safe(d, key, default='-'):
    v = d.get(key)
    return str(v) if v is not None else default   # keeps 0 / False / '0'


def _fmt_date(val) -> str:
    """Return date as 'dd/mm/yyyy' string — never a datetime object.

    Writing a bare datetime via openpyxl produces an Excel serial number unless
    the target cell already carries a date number-format (which our templates
    don't guarantee).  Always use this for date cells.
    """
    if isinstance(val, (datetime, date)):
        return val.strftime('%d/%m/%Y')
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(str(val), fmt).strftime('%d/%m/%Y')
        except (ValueError, TypeError):
            pass
    if val in (None, '', '-'):
        return '-'
    return str(val)


def _combine_fields(*parts, sep='\n') -> str:
    """Join non-empty / non-dash parts. Returns '-' when nothing to combine."""
    filtered = [str(p) for p in parts if p and str(p) not in ('-', '')]
    return sep.join(filtered) if filtered else '-'


def _rich_bold_labels(text) -> object:
    """Return a CellRichText where the portion before ':' on each line is bold.

    Used for multi-line fields like no_of_spans where each line has the form
    "Side Label: value". Returns None if no content is given (blank cell).
    Falls back to plain string if openpyxl rich-text is unavailable.
    """
    if not text or str(text).strip() in ('-', '', 'None'):
        return None
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
    except ImportError:
        return str(text)   # older openpyxl — plain-string fallback

    bold = InlineFont(b=True)
    lines = str(text).split('\n')
    parts = CellRichText()
    for i, line in enumerate(lines):
        newline = '\n' if i < len(lines) - 1 else ''
        if ':' in line:
            label, rest = line.split(':', 1)
            parts.append(TextBlock(bold, label + ':'))
            parts.append(rest + newline)
        else:
            parts.append(line + newline)
    return parts


def _build_spans_cell(d) -> object:
    """Build the 'Number of Spans' cell with 4 fixed bold sub-titles.

    Format (each sub-title bold; omitted when no data for that sub-field):
        Number of Span: <no_of_spans>
        Length:         <total_length>
        C/C of Piers:   <cc_of_piers>
        Width of Piers: <width_of_piers>

    Fallback: if none of the structured sub-fields are available, writes a
    plain text dump of whatever span info was given (no bold formatting).
    Returns None (blank cell) when no data at all is provided.
    """
    def _v(key):
        val = (d.get(key) or '').strip()
        return val if val and val not in ('-', 'None') else ''

    no_of_spans    = _v('no_of_spans')
    total_length   = _v('total_length')
    cc_of_piers    = _v('cc_of_piers')
    width_of_piers = _v('width_of_piers')
    span_length    = _v('span_length')   # legacy fallback

    # Build ordered sections — only include sub-titles that have data
    sections = []
    if no_of_spans:
        sections.append(('Number of Span', no_of_spans))
    if total_length:
        sections.append(('Length', total_length))
    if cc_of_piers:
        sections.append(('C/C of Piers', cc_of_piers))
    if width_of_piers:
        sections.append(('Width of Piers', width_of_piers))

    if not sections:
        # No structured data — fall back to plain text dump
        fallback = _combine_fields(no_of_spans, span_length)
        return _rich_bold_labels(fallback)   # returns None if still empty

    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
    except ImportError:
        # Older openpyxl — plain-string fallback (no bold)
        return '\n'.join(f'{label}: {val}' for label, val in sections)

    bold  = InlineFont(b=True)
    parts = CellRichText()
    for i, (label, value) in enumerate(sections):
        newline = '\n' if i < len(sections) - 1 else ''
        parts.append(TextBlock(bold, label + ':'))
        # Value may itself be multi-line — preserve as-is
        parts.append(' ' + value + newline)
    return parts


def _safe_write(ws, row: int, col: int, value):
    """Write to a cell, unmerging its range first if it is a MergedCell."""
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import range_boundaries
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for rng in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(str(rng))
            if min_row <= row <= max_row and min_col <= col <= max_col:
                ws.unmerge_cells(str(rng))
                break
    ws.cell(row=row, column=col).value = value

def _cell(ws, addr, value):
    """Write to a named cell address (e.g. 'C4'), unmerging if needed."""
    from openpyxl.utils import coordinate_to_tuple
    row, col = coordinate_to_tuple(addr)
    _safe_write(ws, row, col, value)


def _find_sheet(wb, keyword):
    """Find the first sheet whose name contains keyword (case-insensitive)."""
    kw = keyword.lower()
    for name in wb.sheetnames:
        if kw in name.lower():
            return wb[name]
    return None


def _detect_remarks_col(ws):
    """Scan row 3 to find the 1-based column index of the Remarks column."""
    for cell in ws[3]:
        if cell.value and 'remark' in str(cell.value).lower():
            return cell.column
    # Fallback: one past the last non-empty cell in row 3
    last = 4
    for cell in ws[3]:
        if cell.value is not None and cell.column >= 5:
            last = cell.column
    return last + 1


def _fill_defect_table(ws, elements: list, matrix: dict, start_col: int, remarks_col: int, row_offset: int = 0):
    """Fill one defect table sheet.

    ALWAYS clears the data region first to remove stale template data.
    Writes blank instead of 'Absent' for unobserved defects.
    row_offset: extra rows to add to each defect row (use 1 for AMC SUPER sheets
    which have an extra 'Type of Super Structure' row at row 4).
    """
    # Always clear data region — including remarks column (stale photo refs from template)
    clear_end = (start_col + len(elements)) if elements else remarks_col
    for col_i in range(start_col, clear_end):
        _safe_write(ws, 3, col_i, None)
        for row in range(4 + row_offset, 15 + row_offset):
            _safe_write(ws, row, col_i, None)
    # Clear remarks column data rows (keep row-3 header label)
    for row in range(4 + row_offset, 15 + row_offset):
        _safe_write(ws, row, remarks_col, None)

    if not elements:
        return

    for i, elem_id in enumerate(elements):
        _safe_write(ws, 3, start_col + i, elem_id)

    for defect_key, row_num in DEFECT_ROW.items():
        for i, elem_id in enumerate(elements):
            obs = (matrix.get(elem_id, {}) or {}).get(defect_key, '')
            if obs and str(obs).strip().lower() != 'absent':
                _safe_write(ws, row_num + row_offset, start_col + i, obs)


def _fill_defect_tables(wb, d):
    """Find and fill all four defect sheets in the AMC template."""
    START_COL = 5  # column E — same in both R&B and AMC templates

    # Sub-structure Side 1  (sheet renamed from "GURUJI SUB - 1" → "Table 1")
    ws1 = _find_sheet(wb, 'table 1')
    if ws1:
        piers1  = d.get('sub_piers_side1') or []
        matrix1 = d.get('defect_sub1') or {}
        remarks_col = _detect_remarks_col(ws1)
        _fill_defect_table(ws1, piers1, matrix1, START_COL, remarks_col)

    # Sub-structure Side 2  (sheet renamed → "Table 2")
    ws2 = _find_sheet(wb, 'table 2')
    if ws2:
        piers2  = d.get('sub_piers_side2') or []
        matrix2 = d.get('defect_sub2') or {}
        remarks_col = _detect_remarks_col(ws2)
        _fill_defect_table(ws2, piers2, matrix2, START_COL, remarks_col)

    # Super-structure Side 1  (sheet renamed → "Table 3")
    ws3 = _find_sheet(wb, 'table 3')
    if ws3:
        spans1  = d.get('super_spans_side1') or []
        matrix3 = d.get('defect_super1') or {}
        remarks_col = _detect_remarks_col(ws3)
        _fill_defect_table(ws3, spans1, matrix3, START_COL, remarks_col, row_offset=1)

    # Super-structure Side 2  (sheet renamed → "Table 4")
    ws4 = _find_sheet(wb, 'table 4')
    if ws4:
        spans2  = d.get('super_spans_side2') or []
        matrix4 = d.get('defect_super2') or {}
        remarks_col = _detect_remarks_col(ws4)
        _fill_defect_table(ws4, spans2, matrix4, START_COL, remarks_col, row_offset=1)


def _fill_title_page(wb, d):
    ws = wb['TITLE PAGE']
    _cell(ws, 'A2', f"CLIENT: {_safe(d, 'client_name', 'AHMEDABAD MUNICIPAL CORPORATION')}")
    _cell(ws, 'C4', _safe(d, 'project_name', 'Bridge Inspection Work Ahmedabad City'))
    _cell(ws, 'C5', f'"{_safe(d, "bridge_title_full", _safe(d, "bridge_title", d.get("river_name", "")))}"')
    _cell(ws, 'A6', f"Project No.: {_safe(d, 'project_number', '-')}")
    _cell(ws, 'A10', 'R0')
    _cell(ws, 'B10', date.today())
    _cell(ws, 'D10', 'Preliminary Inspection Report')
    _cell(ws, 'E10', 'CASAD')


def _fill_appendix_a(wb, d):
    """Fill Appendix-A — purely user-input driven; no code-generated defaults.
    Each cell is written only when the user explicitly provided a value.
    """
    from openpyxl.cell.cell import MergedCell
    ws = wb['Appendix-A']

    # Clear ALL variable data cells in column C (rows 4–104)
    for r in range(4, 105):
        cell = ws.cell(row=r, column=3)
        if not isinstance(cell, MergedCell):
            cell.value = None

    def _v(key):
        v = d.get(key)
        return v if v and str(v).strip() not in ('-', '') else None

    lat = _v('latitude')
    lon = _v('longitude')
    location = f"{lat}° , {lon}°" if lat and lon else None

    mapping = {
        # ── Section 1 — General identity (rows 4-12) ─────────────────────────
        'C4':  _v('bridge_title') or _v('river_name'),
        'C5':  _v('bridge_number'),
        'C6':  _v('river_name'),
        'C7':  _v('road_name'),
        'C8':  _v('road_number'),
        'C9':  location,
        'C10': _v('bm_gts_level'),
        'C11': _v('division') or _v('circle'),
        'C12': _v('circle'),

        # ── Section 2 — Details of Spans (rows 15-18) ────────────────────────
        'C15': _build_spans_cell(d),
        'C16': _v('total_length'),
        'C17': _v('angle_of_crossing'),
        'C18': _v('bridge_level_type') or _v('type_of_bridge'),

        # ── Section 3 — Hydraulic Parameters (row 20 header + sub-rows 21–31) ─
        'C20': _v('hydraulic_parameters'),
        'C21': _v('hydraulic_catchment'),
        'C22': _v('hydraulic_discharge'),
        'C23': _v('hydraulic_hfl'),
        'C24': _v('hydraulic_ofl'),
        'C25': _v('hydraulic_clearance'),
        'C26': _v('hydraulic_lwl'),
        'C27': _v('hydraulic_depth'),
        'C28': _v('hydraulic_velocity'),
        'C29': _v('hydraulic_channel_width'),
        'C30': _v('hydraulic_spread'),
        'C31': _v('hydraulic_bed_level'),

        # ── Section 4 — Sub Soil Particulars (row 33 header + sub-rows 34–39) ─
        'C33': _v('subsoil_particulars'),
        'C34': _v('subsoil_type'),
        'C35': _v('subsoil_friction'),
        'C36': _v('subsoil_cohesion'),
        'C37': _v('subsoil_silt_factor'),
        'C38': _v('subsoil_bearing_capacity'),
        'C39': _v('subsoil_foundation_level'),

        # ── Section 5 — Design and Structural Data ────────────────────────────
        'C42': _v('loading_standard'),
        'C43': _v('superstructure_type') or _v('bridge_type'),
        'C44': _v('span_arrangement') or _v('span_length'),
        'C45': _v('carriage_width'),
        'C46': _v('deck_level'),
        'C48': _v('design_scour_level'),
        'C49': _v('design_foundation_level'),
        'C50': _v('foundation_type'),
        'C51': _v('substructure_type'),
        'C52': _v('substructure_material'),
        'C53': _v('pier_length'),
        'C54': _v('pier_width_detail'),
        'C55': _v('pier_cap_width'),
        'C56': _v('abutment_width'),
        'C57': _v('abutment_cap_width'),
        'C58': _v('returns_length'),
        'C59': _v('superstructure_type'),
        'C60': _v('prestressing_details'),
        'C61': _v('articulation_details'),
        'C62': _v('total_load_foundation'),
        'C63': _v('total_horizontal_force'),
        'C64': _v('bearing_type_detail'),
        'C65': _v('wearing_coat'),
        'C66': _v('railing_type'),
        'C67': _v('expansion_joint'),
        'C68': _v('protection_works'),
        'C69': _v('model_studies'),
        'C70': _v('special_design_features'),
        'C71': _v('settlement_report'),

        # ── Material Consumed (rows 74–77) ────────────────────────────────────
        'C74': _v('material_cement'),
        'C75': _v('material_reinforcement'),
        'C76': _v('material_structural_steel'),
        'C77': _v('material_hts_steel'),

        # ── Other Data (rows 80–91) ───────────────────────────────────────────
        'C80': _fmt_date(d.get('date_of_construction_start')) if d.get('date_of_construction_start') else None,
        'C81': _fmt_date(d.get('date_of_completion')) if d.get('date_of_completion') else None,
        'C82': _v('surface_utilities'),
        'C83': _v('design_drawings'),
        'C84': _v('ls_sketch'),
        'C85': _v('special_features'),
        'C86': _v('total_cost'),
        'C87': _v('cost_per_sqm_carriageway'),
        'C88': _v('cost_per_sqm_elevation'),
        'C89': _v('cost_per_m_length'),
        'C90': _v('design_agency'),
        'C91': _v('construction_agency'),

        # ── Performance & survey date ─────────────────────────────────────────
        'C92': _v('performance'),
        'C93': _fmt_date(d.get('date_of_survey')) if d.get('date_of_survey') else None,
    }
    for addr, val in mapping.items():
        try:
            _cell(ws, addr, val)
        except Exception:
            pass


def _fill_appendix_b(wb, d):
    """Fill Appendix-B — purely user-input driven.
    Pre-wipe covers the full data area (rows 4-128) so no stale template data bleeds through.
    No hardcoded cross-references or defaults — every cell is written only from user input.
    Section 20 (recommendations) is intentionally left blank for the engineer.
    """
    from openpyxl.cell.cell import MergedCell
    ws = wb['Appendix-B']

    # Wipe full data area (rows 4-128, col C)
    for r in range(4, 129):
        cell = ws.cell(row=r, column=3)
        if not isinstance(cell, MergedCell):
            cell.value = None

    def _v(key):
        v = d.get(key)
        return v if v and str(v).strip() not in ('-', '') else None

    lat = _v('latitude')
    lon = _v('longitude')
    location = f"{lat}° , {lon}°" if lat and lon else None

    div = _v('division')
    cir = _v('circle')
    if div and cir and div != cir:
        div_cir = f"{div} / {cir}"
    else:
        div_cir = div or cir or None

    fields = {
        # ── Section 1 — General ───────────────────────────────────────────────
        'C4':  _v('bridge_title') or _v('river_name'),
        'C5':  _v('bridge_number'),
        'C6':  _v('river_name'),
        'C7':  _v('road_name'),
        'C8':  _v('road_number'),
        'C9':  location,
        'C10': div_cir,
        'C11': _v('type_of_bridge') or _v('bridge_type'),
        'C12': _fmt_date(d.get('date_of_survey')) if d.get('date_of_survey') else None,

        # ── Section 4 — Approaches ────────────────────────────────────────────
        'C14': _v('approach_settlement'),
        'C15': _v('approach_side_slopes'),
        'C16': _v('approach_erosion'),
        'C17': _v('approach_slab'),
        'C18': _v('approach_geometrics'),
        'C19': _v('approach_other'),

        # ── Section 5 — Protective Works ─────────────────────────────────────
        'C21': _v('prot_type'),
        'C22': _v('prot_damage_layout'),
        'C23': _v('prot_slope_pitching'),
        'C24': _v('prot_floor_protection'),
        'C25': _v('prot_scour_extent'),
        'C26': _v('prot_reserve_stone'),
        'C27': _v('prot_other'),

        # ── Section 6 — Waterway ──────────────────────────────────────────────
        'C28': _v('waterway_obs'),
        'C29': _v('waterway_obstruction'),
        'C30': _v('waterway_scour'),
        'C31': _v('waterway_flow'),
        'C32': _v('waterway_flood_level'),
        'C33': _v('waterway_afflux'),
        'C34': _v('waterway_adequacy'),
        'C35': _v('waterway_other'),

        # ── Section 7 — Foundations ───────────────────────────────────────────
        'C36': _v('foundations_obs'),
        'C37': _v('foundations_settlement'),
        'C38': _v('foundations_cracking'),
        'C39': _v('foundations_floating'),
        'C40': _v('foundations_subway'),
        'C41': _v('foundations_other'),

        # ── Section 8 — Substructure (no hardcoded cross-refs) ───────────────
        'C42': _v('sub_section_obs'),
        'C43': _v('sub_drainage_backfill'),
        'C44': _v('sub_cracking_obs'),
        'C45': _v('sub_subway_obs'),
        'C46': _v('sub_other_obs'),

        # ── Section 9 — Bearings ──────────────────────────────────────────────
        'C48': _v('bear_metallic_type'),
        'C49': _v('bear_metallic_condition'),
        'C50': _v('bear_metallic_functioning'),
        'C51': _v('bear_metallic_greasing'),
        'C52': _v('bear_pedestal_cracks'),
        'C53': _v('bear_metallic_anchor'),
        'C54': _v('bear_metallic_other'),
        'C55': _v('bear_elastomeric_type'),
        'C56': _v('bear_pad_condition'),
        'C57': _v('bear_cleanliness'),
        'C58': _v('bear_elastomeric_other'),

        # ── Section 10 — Superstructure (no hardcoded cross-refs) ────────────
        'C59': _v('super_section_obs'),
        'C60': _v('superstructure_type'),
        'C61': _v('super_spalling_obs'),
        'C62': _v('super_cracking_obs'),
        'C63': _v('super_corrosion_obs'),
        'C64': _v('super_vehicle_damage'),
        'C65': _v('super_articulation'),
        'C66': _v('super_vibration'),
        'C67': _v('super_deflection'),
        'C68': _v('super_anchorage_cracks'),
        'C69': _v('super_hinge_deflection'),
        'C70': _v('steel_obs'),
        'C71': _v('steel_paint'),
        'C72': _v('steel_corrosion'),
        'C73': _v('steel_vibration'),
        'C74': _v('steel_alignment'),
        'C75': _v('steel_connections'),
        'C76': _v('steel_camber_deflection'),
        'C77': _v('steel_buckling'),
        'C78': _v('steel_cleanliness'),
        'C79': _v('masonry_obs'),
        'C80': _v('masonry_joints'),
        'C81': _v('masonry_profile'),
        'C82': _v('masonry_cracks'),
        'C83': _v('masonry_drainage'),
        'C84': _v('masonry_vegetation'),
        'C85': _v('masonry_other'),
        'C86': _v('timber_obs'),
        'C87': _v('timber_paint'),
        'C88': _v('timber_decay'),
        'C89': _v('timber_joints'),
        'C90': _v('timber_sag'),

        # ── Section 11 — Expansion Joints ────────────────────────────────────
        'C92': _v('exp_jt_functioning'),
        'C93': _v('exp_jt_sealing'),
        'C94': _v('exp_jt_fixing'),
        'C95': _v('exp_jt_sliding_plate'),
        'C96': _v('exp_jt_locking'),
        'C97': _v('exp_jt_debris'),
        'C98': _v('exp_jt_rattling'),
        'C99': _v('exp_jt_other'),

        # ── Section 12 — Wearing Coat ─────────────────────────────────────────
        'C100': _v('wear_coat_type'),
        'C101': _v('wear_coat_surface'),
        'C102': _v('wear_coat_evidence'),

        # ── Section 13 — Drainage Spouts ─────────────────────────────────────
        'C103': _v('drain_type'),
        'C104': _v('drain_clogging'),
        'C105': _v('drain_projection'),
        'C106': _v('drain_adequacy'),
        'C107': _v('drain_subway'),
        'C108': _v('drain_other'),

        # ── Section 14 — Handrail ─────────────────────────────────────────────
        'C110': _v('handrail_condition'),
        'C111': _v('handrail_collision'),
        'C112': _v('handrail_alignment'),

        # ── Section 15 — Footpath ─────────────────────────────────────────────
        'C114': _v('footpath_condition'),
        'C115': _v('footpath_missing_slab'),
        'C116': _v('footpath_other'),

        # ── Section 16 — Utilities ────────────────────────────────────────────
        'C117': _v('utilities_obs'),
        'C118': _v('util_water_leakage'),
        'C119': _v('util_cable_damage'),
        'C120': _v('util_lighting'),
        'C121': _v('util_other_damage'),

        # ── Section 17 — Bridge Number ────────────────────────────────────────
        'C123': _v('bridge_num_condition'),

        # ── Section 18 — Aesthetics ───────────────────────────────────────────
        'C125': _v('aesthetics_intrusion'),

        # ── Section 19 — Maintenance History ─────────────────────────────────
        'C126': _v('maintenance_history'),

        # ── Overall Condition (unlabelled row 127) ────────────────────────────
        'C127': _v('overall_condition_visual'),

        # ── Section 20 — Recommendations: intentionally not written ──────────
        # Engineer fills this cell manually.
    }
    for addr, val in fields.items():
        try:
            _cell(ws, addr, val)
        except Exception:
            pass


def _has_red_markers(path: str) -> bool:
    """Return True if the image already has prominent hand-drawn red circles (2% threshold)."""
    try:
        from PIL import Image as _PIL
        img = _PIL.open(path).convert('RGB').resize((120, 120))
        px  = list(img.getdata())
        red = sum(1 for r, g, b in px if r > 170 and g < 90 and b < 90)
        return red / len(px) > 0.02
    except Exception:
        return False


def _fill_appendix_c(wb, d):
    """Insert photos into the single Appendix-C- sheet (AMC format).

    Photos are inserted WITHOUT burning circles — editable ovals are injected
    as Excel AutoShapes after save.  Returns oval descriptor list.
    """
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor

    ws = _find_sheet(wb, 'appendix - photos')
    if ws is None:
        print("AMC PHOTO: Appendix - Photos sheet not found, skipping photos.")
        return []

    photos      = d.get('photos', [])
    titles      = d.get('photo_titles', [])
    categories  = d.get('photo_categories', [])
    coords_list = list(d.get('photo_coords', [])) + [None] * len(photos)

    # Pad lists to equal length
    max_len     = max(len(photos), len(titles), len(categories)) if any([photos, titles, categories]) else 0
    photos      = (photos      + [''] * max_len)[:max_len]
    titles      = (titles      + [''] * max_len)[:max_len]
    categories  = (categories  + [''] * max_len)[:max_len]
    coords_list = coords_list[:max_len]

    CAPTION_FILL   = PatternFill(patternType='solid', fgColor='FCE4D6')
    CAPTION_FONT   = Font(name='Times New Roman', size=11, bold=True)
    CAPTION_ALIGN  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    _thin          = Side(style='thin', color='000000')
    CAPTION_BORDER = Border(top=_thin, left=_thin, right=_thin, bottom=_thin)

    ws._images.clear()
    for rng in list(ws.merged_cells.ranges):
        r = str(rng)
        if 'B2' not in r and 'A1' not in r:
            try:
                ws.unmerge_cells(r)
            except Exception:
                pass

    ovals     = []
    shape_ctr = 200   # start higher than R&B module to avoid ID collision

    row = 3
    for path, title, cat, coords in zip(photos, titles, categories, coords_list):
        if not path or not os.path.exists(path):
            continue
        try:
            from PIL import Image as PILImage
            with PILImage.open(path) as img:
                img.load()
                if img.mode in ('RGBA', 'P', 'LA'):
                    img = img.convert('RGB')
                # No burned-in circle — editable oval injected after save
                w, h    = img.size
                scale   = min(1.0, 1200 / w, 900 / h)
                new_w   = int(w * scale)
                new_h   = int(h * scale)
                buf     = io.BytesIO()
                img.resize((new_w, new_h), PILImage.LANCZOS).save(buf, format='JPEG', quality=90)
            buf.seek(0)

            ph_from_row = row - 1
            # Approximate rows occupied: default row height ~15pt ≈ 20px at 96 dpi
            import math as _math
            rows_occupied = max(10, _math.ceil(new_h / 20))
            ph_to_row     = ph_from_row + rows_occupied

            xl_img        = XLImage(buf)
            xl_img.width  = new_w   # OneCellAnchor respects these dimensions — no stretching
            xl_img.height = new_h
            anchor        = OneCellAnchor()
            anchor._from  = AnchorMarker(col=0, row=ph_from_row, colOff=0, rowOff=0)
            xl_img.anchor = anchor
            ws.add_image(xl_img)

            # Schedule editable oval if defect coords available
            if coords and not _has_red_markers(path):
                x_pct, y_pct = coords
                span_cols = 8
                span_rows = ph_to_row - ph_from_row   # 20
                r_cols = max(1, int(span_cols * 0.07))
                r_rows = max(1, int(span_rows * 0.07))
                oval_fc = max(0, int(x_pct * span_cols) - r_cols)
                oval_tc = min(span_cols, int(x_pct * span_cols) + r_cols)
                oval_fr = ph_from_row + max(0, int(y_pct * span_rows) - r_rows)
                oval_tr = ph_from_row + min(span_rows, int(y_pct * span_rows) + r_rows)
                shape_ctr += 1
                ovals.append((ws.title, oval_fc, oval_fr, oval_tc, oval_tr, shape_ctr))

            cap_row = ph_to_row + 2
            try:
                ws.merge_cells(start_row=cap_row, start_column=1,
                               end_row=cap_row, end_column=8)
            except Exception:
                pass
            cap_cell           = ws.cell(row=cap_row, column=1, value=title or path)
            cap_cell.fill      = CAPTION_FILL
            cap_cell.font      = CAPTION_FONT
            cap_cell.alignment = CAPTION_ALIGN
            cap_cell.border    = CAPTION_BORDER
            ws.row_dimensions[cap_row].height = 28

            row = cap_row + 3

        except Exception as e:
            print(f"AMC PHOTO INSERT FAILED {path}: {e}")

    return ovals


def build_excel_amc(report_json: dict) -> str:
    """Fill CASAD AMC Excel template with report_json and return saved file path."""
    if _AMC_TEMPLATE_BYTES:
        wb = openpyxl.load_workbook(io.BytesIO(_AMC_TEMPLATE_BYTES))
    else:
        wb = openpyxl.load_workbook(AMC_TEMPLATE_PATH)
    try:
        _fill_title_page(wb, report_json)
        _fill_appendix_a(wb, report_json)
        _fill_appendix_b(wb, report_json)
        _fill_defect_tables(wb, report_json)
        ovals = _fill_appendix_c(wb, report_json)

        bridge   = re.sub(r'[^\w\-]', '_', report_json.get('bridge_title', report_json.get('river_name', 'bridge')))
        date_str = report_json.get('date_of_survey', 'report').replace('/', '-')
        out_path = os.path.join(OUTPUT_DIR, f'CASAD_AMC_{bridge}_{date_str}.xlsx')
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        wb.save(out_path)
        print(f"AMC EXCEL REPORT SAVED: {out_path}")
    finally:
        wb.close()

    # Inject editable oval shapes after save
    if ovals:
        try:
            from report_gen_excel import _inject_oval_shapes
            _inject_oval_shapes(out_path, ovals)
        except Exception as e:
            print(f"AMC OVAL INJECT FAILED: {e}")

    return out_path
