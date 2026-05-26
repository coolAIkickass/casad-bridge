# report_gen_excel.py — Fill CASAD Excel template from report JSON
import os, re, io
from datetime import datetime, date
import openpyxl
from openpyxl.drawing.image import Image as XLImage

EXCEL_TEMPLATE_PATH = os.getenv('EXCEL_TEMPLATE_PATH', 'casad_excel_template.xlsx')
OUTPUT_DIR = os.getenv('OUTPUT_DIR', 'media')

# Cache raw template bytes at import time — avoids repeated disk reads under load.
# Each build_excel() call loads from memory via io.BytesIO.
try:
    with open(EXCEL_TEMPLATE_PATH, 'rb') as _f:
        _TEMPLATE_BYTES = _f.read()
except FileNotFoundError:
    _TEMPLATE_BYTES = None

DEFECT_ROW = {
    'cracks': 4, 'leaching': 5, 'honeycombing': 6, 'exposed_rebar': 7,
    'leakage': 8, 'spalling': 9, 'rust_marks': 10, 'shuttering': 11,
    'delamination': 12, 'vegetation': 13, 'any_other': 14
}

# ─────────────────────────────────────────────────────────────
#  Shared helpers
# ─────────────────────────────────────────────────────────────

def _safe(d, key, default='-'):
    """Return str(value) or default.  Treats None as missing; keeps 0/'0'/False."""
    v = d.get(key)
    return str(v) if v is not None else default


def _parse_survey_date(val):
    """Parse 'dd/mm/yyyy' string to a datetime object (internal use only)."""
    if isinstance(val, (datetime, date)):
        return val
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(str(val), fmt)
        except (ValueError, TypeError):
            pass
    return val   # fallback: return as-is


def _fmt_date(val) -> str:
    """Return date as 'dd/mm/yyyy' string — NEVER a datetime object.

    openpyxl writes a bare datetime as an Excel serial number unless the cell
    already has a date number-format applied (which our templates don't
    guarantee).  Always use this function for date cells in the report.
    """
    dt = _parse_survey_date(val)
    if isinstance(dt, (datetime, date)):
        return dt.strftime('%d/%m/%Y')
    if val in (None, '', '-'):
        return '-'
    return str(val)


def _combine_fields(*parts, sep='\n') -> str:
    """Join non-empty / non-dash parts with *sep*.  Returns '-' if nothing."""
    filtered = [str(p) for p in parts if p and str(p) not in ('-', '')]
    return sep.join(filtered) if filtered else '-'


def _rich_bold_labels(text) -> object:
    """Return a CellRichText where the portion before ':' on each line is bold.

    Used for multi-line fields like no_of_spans where each line has the form
    "Side Label: value". Returns None if no content is given (blank cell).
    Falls back to plain string if openpyxl rich-text is unavailable.
    """
    if not text or str(text).strip() in ('-', '', 'None'):
        return None
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
    except ImportError:
        return str(text)   # older openpyxl — plain-string fallback

    bold = InlineFont(b=True)
    lines = str(text).split('\n')
    parts = CellRichText()
    for i, line in enumerate(lines):
        newline = '\n' if i < len(lines) - 1 else ''
        if ':' in line:
            label, rest = line.split(':', 1)
            parts.append(TextBlock(bold, label + ':'))
            parts.append(rest + newline)
        else:
            parts.append(line + newline)
    return parts


def _build_spans_cell(d) -> object:
    """Build the 'Number of Spans' cell — only the span count/arrangement.

    Other fields (total_length, cc_of_piers, width_of_piers) each have their
    own dedicated rows in the template and must not be duplicated here.
    Returns None (blank cell) when no span data is provided.
    """
    val = (d.get('no_of_spans') or '').strip()
    if not val or val in ('-', 'None'):
        return None
    return _rich_bold_labels('Number of Span: ' + val)


def _coords(d) -> str:
    """Format lat/lon with degree symbol."""
    lat = d.get('latitude', '-')
    lon = d.get('longitude', '-')
    return f"{lat}° , {lon}°"


def _safe_write(ws, row: int, col: int, value):
    """Write to a cell, unmerging its range first if it is a MergedCell."""
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import range_boundaries
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for rng in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(str(rng))
            if min_row <= row <= max_row and min_col <= col <= max_col:
                ws.unmerge_cells(str(rng))
                break
    ws.cell(row=row, column=col).value = value


def _cell(ws, addr, value):
    """Write to a named cell address (e.g. 'C4'), unmerging if needed."""
    from openpyxl.utils import coordinate_to_tuple
    row, col = coordinate_to_tuple(addr)
    _safe_write(ws, row, col, value)


def _merge_write(ws, row1: int, col1: int, row2: int, col2: int, value):
    """Unmerge anything in the target area, write value to top-left, re-merge."""
    # Unmerge every existing range overlapping this block
    from openpyxl.utils import range_boundaries
    for rng in list(ws.merged_cells.ranges):
        rc1, rr1, rc2, rr2 = range_boundaries(str(rng))
        if rr1 <= row2 and rr2 >= row1 and rc1 <= col2 and rc2 >= col1:
            ws.unmerge_cells(str(rng))
    ws.cell(row=row1, column=col1).value = value
    if row1 != row2 or col1 != col2:
        ws.merge_cells(start_row=row1, start_column=col1,
                       end_row=row2, end_column=col2)


# ─────────────────────────────────────────────────────────────
#  Sheet fillers
# ─────────────────────────────────────────────────────────────

def _fill_title_page(wb, d):
    ws = wb['TITLE PAGE']
    _cell(ws, 'A2', f"CLIENT: {_safe(d, 'client_name', 'CASAD CONSULTANTS PVT. LTD.')}")
    _cell(ws, 'C4', _safe(d, 'project_name', 'Bridge Inspection Work'))
    _cell(ws, 'C5', f'"{_safe(d, "bridge_title_full", _safe(d, "bridge_title", d.get("river_name", "")))}"')
    _cell(ws, 'A6', f"Project No.: {_safe(d, 'project_number', '-')}")
    # Revision row — use report date, not today; revision number from fixture or default R0
    rev_no   = _safe(d, 'revision_number', 'R0')
    rev_date = _parse_survey_date(d.get('date_of_survey', date.today()))
    _cell(ws, 'A10', rev_no)
    _cell(ws, 'B10', rev_date)
    _cell(ws, 'D10', 'Preliminary Inspection Report')
    _cell(ws, 'E10', 'CASAD')


def _fill_appendix_a(wb, d):
    ws = wb['Appendix-A']

    # ── Step 1: wipe ALL variable data cells (column C, rows 4-104) ──────────
    # This prevents stale data from a previous bridge report leaking into the
    # new one (the template is reused and may carry old values).
    from openpyxl.cell.cell import MergedCell
    for r in range(4, 105):
        cell = ws.cell(row=r, column=3)   # col C = 3
        if not isinstance(cell, MergedCell):
            cell.value = None

    # ── Step 2: write current report values ──────────────────────────────────
    def _v(key):
        v = d.get(key)
        return v if v and str(v).strip() not in ('-', '') else None

    lat = _v('latitude')
    lon = _v('longitude')
    location = f"{lat}° , {lon}°" if lat and lon else None

    mapping = {
        # Section 1 — General identity
        'C4':  _v('bridge_title') or _v('river_name'),
        'C5':  _v('bridge_number'),
        'C6':  _v('river_name'),
        'C7':  _v('road_name'),
        'C8':  _v('road_number'),
        'C9':  location,
        'C10': _v('bm_gts_level'),
        'C11': _v('division') or _v('circle'),
        'C12': _v('circle'),

        # Section 2 — Details of Spans
        'C15': _build_spans_cell(d),
        'C16': _v('total_length'),
        'C17': _v('angle_of_crossing'),
        'C18': _v('bridge_level_type') or _v('type_of_bridge'),

        # Section 3 — Hydraulic Parameters (header + sub-rows 21–31)
        'C20': _v('hydraulic_parameters'),
        'C21': _v('hydraulic_catchment'),
        'C22': _v('hydraulic_discharge'),
        'C23': _v('hydraulic_hfl'),
        'C24': _v('hydraulic_ofl'),
        'C25': _v('hydraulic_clearance'),
        'C26': _v('hydraulic_lwl'),
        'C27': _v('hydraulic_depth'),
        'C28': _v('hydraulic_velocity'),
        'C29': _v('hydraulic_channel_width'),
        'C30': _v('hydraulic_spread'),
        'C31': _v('hydraulic_bed_level'),

        # Section 4 — Sub Soil Particulars (header + sub-rows 34–39)
        'C33': _v('subsoil_particulars'),
        'C34': _v('subsoil_type'),
        'C35': _v('subsoil_friction'),
        'C36': _v('subsoil_cohesion'),
        'C37': _v('subsoil_silt_factor'),
        'C38': _v('subsoil_bearing_capacity'),
        'C39': _v('subsoil_foundation_level'),

        # Section 5 — Design and Structural Data
        'C42': _v('loading_standard'),
        'C43': _v('superstructure_type') or _v('bridge_type'),
        'C44': _v('span_arrangement') or _v('span_length'),
        'C45': _v('carriage_width'),
        'C46': _v('deck_level'),
        'C48': _v('design_scour_level'),
        'C49': _v('design_foundation_level'),
        'C50': _v('foundation_type'),
        'C51': _v('substructure_type'),
        'C52': _v('substructure_material'),
        'C53': _v('pier_length'),
        'C54': _v('pier_width_detail'),
        'C55': _v('pier_cap_width'),
        'C56': _v('abutment_width'),
        'C57': _v('abutment_cap_width'),
        'C58': _v('returns_length'),
        'C59': _v('superstructure_type'),
        'C60': _v('prestressing_details'),
        'C61': _v('articulation_details'),
        'C62': _v('total_load_foundation'),
        'C63': _v('total_horizontal_force'),
        'C64': _v('bearing_type_detail'),
        'C65': _v('wearing_coat'),
        'C66': _v('railing_type'),
        'C67': _v('expansion_joint'),
        'C68': _v('protection_works'),
        'C69': _v('model_studies'),
        'C70': _v('special_design_features'),
        'C71': _v('settlement_report'),

        # Material Consumed (row 73 header + rows 74–77)
        'C73': _v('material_consumed'),
        'C74': _v('material_cement'),
        'C75': _v('material_reinforcement'),
        'C76': _v('material_structural_steel'),
        'C77': _v('material_hts_steel'),

        # Other Data (rows 80–91)
        'C80': _fmt_date(d.get('date_of_construction_start')) if d.get('date_of_construction_start') else None,
        'C81': _fmt_date(d.get('date_of_completion')) if d.get('date_of_completion') else None,
        'C82': _v('surface_utilities'),
        'C83': _v('design_drawings'),
        'C84': _v('ls_sketch'),
        'C85': _v('special_features'),
        'C86': _v('total_cost'),
        'C87': _v('cost_per_sqm_carriageway'),
        'C88': _v('cost_per_sqm_elevation'),
        'C89': _v('cost_per_m_length'),
        'C90': _v('design_agency'),
        'C91': _v('construction_agency'),

        # Performance & recording date
        'C92': _v('performance'),
        'C93': _fmt_date(d.get('date_of_survey')) if d.get('date_of_survey') else None,
    }
    for addr, val in mapping.items():
        try:
            _cell(ws, addr, val)
        except Exception:
            pass


def _fill_appendix_b(wb, d):
    from openpyxl.cell.cell import MergedCell
    ws = wb['Appendix-B']

    # Pre-wipe col C, rows 4–129
    for r in range(4, 130):
        cell = ws.cell(row=r, column=3)
        if not isinstance(cell, MergedCell):
            cell.value = None

    def _v(key):
        v = d.get(key)
        return v if v and str(v).strip() not in ('-', '') else None

    lat = _v('latitude')
    lon = _v('longitude')
    location = f"{lat}° , {lon}°" if lat and lon else None

    div = _v('division')
    cir = _v('circle')
    div_cir = (f"{div} / {cir}" if div and cir and div != cir else div or cir or None)

    fields = {
        # Section 1 — General identity
        'C4':  _v('bridge_title') or _v('river_name'),
        'C5':  _v('bridge_number'),
        'C6':  _v('river_name'),
        'C7':  _v('road_name'),
        'C8':  _v('road_number'),
        'C9':  location,
        'C10': div_cir,
        'C11': _v('type_of_bridge') or _v('bridge_type'),
        'C12': _fmt_date(d.get('date_of_survey')) if d.get('date_of_survey') else None,

        # Section 4 — Approaches (row 13 = header, data rows 14–19)
        'C14': _v('approach_settlement'),
        'C15': _v('approach_side_slopes'),
        'C16': _v('approach_erosion'),
        'C17': _v('approach_slab'),
        'C18': _v('approach_geometrics'),
        'C19': _v('approach_other'),

        # Section 5 — Protective Works (row 20 = header, data rows 21–27)
        'C21': _v('prot_type'),
        'C22': _v('prot_damage_layout'),
        'C23': _v('prot_slope_pitching'),
        'C24': _v('prot_floor_protection'),
        'C25': _v('prot_scour_extent'),
        'C26': _v('prot_reserve_stone'),
        'C27': _v('prot_other'),

        # Section 6 — Waterway (row 28 = header/summary, data rows 29–35)
        'C28': _v('waterway_obs'),
        'C29': _v('waterway_obstruction'),
        'C30': _v('waterway_scour'),
        'C31': _v('waterway_flow'),
        'C32': _v('waterway_flood_level'),
        'C33': _v('waterway_afflux'),
        'C34': _v('waterway_adequacy'),
        'C35': _v('waterway_other'),

        # Section 7 — Foundations (row 36 = header/summary, data rows 37–41)
        'C36': _v('foundations_obs'),
        'C37': _v('foundations_settlement'),
        'C38': _v('foundations_cracking'),
        'C39': _v('foundations_floating'),
        'C40': _v('foundations_subway'),
        'C41': _v('foundations_other'),

        # Section 8 — Substructure (row 42 = header, data rows 43–46)
        'C42': _v('sub_section_obs'),
        'C43': _v('sub_drainage_backfill'),
        'C44': _v('sub_cracking_obs'),
        'C45': _v('sub_subway_obs'),
        'C46': _v('sub_other_obs'),

        # Section 9 — Bearings (row 47 = header, metallic 48–54, elastomeric 55–58)
        'C48': _v('bear_metallic_type'),
        'C49': _v('bear_metallic_condition'),
        'C50': _v('bear_metallic_functioning'),
        'C51': _v('bear_metallic_greasing'),
        'C52': _v('bear_pedestal_cracks'),
        'C53': _v('bear_metallic_anchor'),
        'C54': _v('bear_metallic_other'),
        'C55': _v('bear_elastomeric_type'),
        'C56': _v('bear_pad_condition'),
        'C57': _v('bear_cleanliness'),
        'C58': _v('bear_elastomeric_other'),

        # Section 10 — Superstructure (row 59 = header, RC/PSC 60–69, steel 70, masonry 79, timber 86)
        'C59': _v('super_section_obs'),
        'C60': _v('superstructure_type'),
        'C61': _v('super_spalling_obs'),
        'C62': _v('super_cracking_obs'),
        'C63': _v('super_corrosion_obs'),
        'C64': _v('super_vehicle_damage'),
        'C65': _v('super_articulation'),
        'C66': _v('super_vibration'),
        'C67': _v('super_deflection'),
        'C68': _v('super_anchorage_cracks'),
        'C69': _v('super_hinge_deflection'),
        'C70': _v('steel_obs'),
        'C79': _v('masonry_obs'),
        'C86': _v('timber_obs'),

        # Section 11 — Expansion Joints (row 91 = header, data rows 92–99)
        'C92': _v('exp_jt_functioning'),
        'C93': _v('exp_jt_sealing'),
        'C94': _v('exp_jt_fixing'),
        'C95': _v('exp_jt_sliding_plate'),
        'C96': _v('exp_jt_locking'),
        'C97': _v('exp_jt_debris'),
        'C98': _v('exp_jt_rattling'),
        'C99': _v('exp_jt_other'),

        # Section 12 — Wearing Coat (row 100 = type/header, 101–102 = sub-rows)
        'C100': _v('wear_coat_type'),
        'C101': _v('wear_coat_surface'),
        'C102': _v('wear_coat_evidence'),

        # Section 13 — Drainage Spouts (row 103 = header, data rows 104–108)
        'C103': _v('drain_type'),
        'C104': _v('drain_clogging'),
        'C105': _v('drain_projection'),
        'C106': _v('drain_adequacy'),
        'C107': _v('drain_subway'),
        'C108': _v('drain_other'),

        # Section 14 — Handrail (row 109 = header, data rows 110–112)
        'C110': _v('handrail_condition'),
        'C111': _v('handrail_collision'),
        'C112': _v('handrail_alignment'),

        # Section 15 — Footpath (row 113 = header, data rows 114–116)
        'C114': _v('footpath_condition'),
        'C115': _v('footpath_missing_slab'),
        'C116': _v('footpath_other'),

        # Section 16 — Utilities (row 117 = header/summary, data rows 118–121)
        'C117': _v('utilities_obs'),
        'C118': _v('util_water_leakage'),
        'C119': _v('util_cable_damage'),
        'C120': _v('util_lighting'),
        'C121': _v('util_other_damage'),

        # Section 17 — Bridge Number (row 122 = header, 123 = observation)
        'C123': _v('bridge_num_condition'),

        # Section 18 — Aesthetics (row 124 = header, 125 = observation)
        'C125': _v('aesthetics_intrusion'),

        # Section 19 — Maintenance history (row 126)
        'C126': _v('maintenance_history'),

        # Row 128 — Overall Condition of Bridge
        'C128': _v('overall_condition_visual'),

        # Row 129 — Section 20 Maintenance & improvement recommendations:
        # intentionally NOT written — engineer fills manually.
    }
    for addr, val in fields.items():
        try:
            _cell(ws, addr, val)
        except Exception:
            pass



# ─────────────────────────────────────────────────────────────
#  Defect tables
# ─────────────────────────────────────────────────────────────

def _fill_defect_table(ws, elements: list, matrix: dict,
                       start_col: int, remarks_col: int):
    """Fill one defect table.

    ALWAYS clears the data region first — prevents stale template data from
    surviving even when 0 inspection data is provided (user gave no defects).

    Writes blank (None) instead of 'Absent' for unobserved defects.
    Only actual observations are written; cells with no data remain blank.
    """
    # Step 1: Always clear — even when elements is empty — to remove old data.
    clear_end = (start_col + len(elements)) if elements else remarks_col
    for col_i in range(start_col, clear_end):
        _safe_write(ws, 3, col_i, None)
        for row in range(4, 15):
            _safe_write(ws, row, col_i, None)

    if not elements:
        return

    # Step 2: Write element IDs in row 3
    for i, elem_id in enumerate(elements):
        _safe_write(ws, 3, start_col + i, elem_id)

    # Step 3: Write only actual observations — leave blank otherwise
    for defect_key, row_num in DEFECT_ROW.items():
        for i, elem_id in enumerate(elements):
            obs = (matrix.get(elem_id, {}) or {}).get(defect_key, '')
            # Skip 'Absent' and empty — leave cell blank (already cleared in step 1)
            if obs and str(obs).strip().lower() != 'absent':
                _safe_write(ws, row_num, start_col + i, obs)


def _fill_defect_tables(wb, d):
    # ── Table 1: Sub-structure Side 1 ──────────────────────────────────────
    # 10 piers → cols E(5)–N(14); Remarks at col O(15)
    piers1  = d.get('sub_piers_side1') or []
    matrix1 = d.get('defect_sub1') or {}
    _fill_defect_table(wb['Table 1'], piers1, matrix1,
                       start_col=5, remarks_col=15)

    # ── Table 2: Sub-structure Side 2 ──────────────────────────────────────
    # Remarks at col O(15)
    piers2  = d.get('sub_piers_side2') or []
    matrix2 = d.get('defect_sub2') or {}
    _fill_defect_table(wb['Table 2'], piers2, matrix2,
                       start_col=5, remarks_col=15)

    # ── Table 3: Super-structure Side 1 ────────────────────────────────────
    # Special layout: Railway span occupies cols E–G (merged), road spans
    # start at col H(8).  Remarks at col S(19).
    spans1  = d.get('super_spans_side1') or []
    matrix3 = d.get('defect_super1') or {}
    ws3     = wb['Table 3']
    if spans1:
        # 1. Clear ONLY the columns we will actually write:
        #    railway in E:G (5–7) + road spans in H.. (8 + n–1).
        #    Do NOT clear beyond the last data column (avoids wiping pre-existing
        #    template headers or merged blocks beyond our data range).
        road_spans     = spans1[1:]
        last_road_col  = 7 + len(road_spans)          # col H=8 → 8 + len-1 = 7+len
        for col_i in range(5, last_road_col + 1):
            for row in range(3, 15):
                _safe_write(ws3, row, col_i, None)

        # 2. Railway span (first element) → merged across E:G
        rly_span    = spans1[0]
        rly_defects = (matrix3.get(rly_span) or {})

        # Header row 3: E3:G3 merged
        _merge_write(ws3, 3, 5, 3, 7, rly_span)

        # Concrete defect rows 4–12: one merged "Not Applicable" block (steel truss)
        concrete_obs = rly_defects.get('concrete', 'Not Applicable')
        _merge_write(ws3, 4, 5, 12, 7, concrete_obs)

        # Row 13 vegetation — merged E13:G13
        veg = rly_defects.get('vegetation') or None
        _merge_write(ws3, 13, 5, 13, 7, veg)

        # Row 14 any_other — merged E14:G14 (corrosion/bolt defects for steel truss).
        # If 'any_other' is absent but 'rust_marks' has content, use rust_marks
        # (the AI occasionally stores corrosion data under rust_marks for steel spans).
        other = (rly_defects.get('any_other')
                 or rly_defects.get('rust_marks')
                 or None)
        _merge_write(ws3, 14, 5, 14, 7, other)

        # 3. Road spans starting at col H(8)
        for i, span_id in enumerate(road_spans):
            col = 8 + i
            _safe_write(ws3, 3, col, span_id)
            for defect_key, row_num in DEFECT_ROW.items():
                obs = (matrix3.get(span_id, {}) or {}).get(defect_key)
                # Write blank (None) for unobserved defects — never write 'Absent'
                val = obs if obs and str(obs).strip().lower() != 'absent' else None
                _safe_write(ws3, row_num, col, val)

    # ── Table 4: Super-structure Side 2 ────────────────────────────────────
    # Remarks at col O(15)
    spans2  = d.get('super_spans_side2') or []
    matrix4 = d.get('defect_super2') or {}
    _fill_defect_table(wb['Table 4'], spans2, matrix4,
                       start_col=5, remarks_col=15)


# ─────────────────────────────────────────────────────────────
#  Photo helpers
# ─────────────────────────────────────────────────────────────

def _has_red_markers(path: str) -> bool:
    """Return True if the image already has prominent hand-drawn red circles.

    Threshold raised to 2 % to avoid false-positives on construction-site
    photos that contain some natural red tones (warning signs, machinery).
    """
    try:
        from PIL import Image as _PIL
        img = _PIL.open(path).convert('RGB').resize((120, 120))
        px  = list(img.getdata())
        red = sum(1 for r, g, b in px if r > 170 and g < 90 and b < 90)
        return red / len(px) > 0.02   # 2 % threshold (was 0.3 %)
    except Exception:
        return False


def _draw_red_circle(img, x_pct: float, y_pct: float):
    """Draw a proportional red circle at the relative defect position."""
    from PIL import ImageDraw
    w, h  = img.size
    cx    = int(x_pct * w)
    cy    = int(y_pct * h)
    r     = max(18, int(min(w, h) * 0.07))
    stroke = max(3,  int(min(w, h) * 0.012))
    ImageDraw.Draw(img).ellipse(
        [cx - r, cy - r, cx + r, cy + r], outline='red', width=stroke)
    return img


# ─────────────────────────────────────────────────────────────
#  Appendix-C photo sheets
# ─────────────────────────────────────────────────────────────

def _find_sheet_rb(wb, name):
    """Return a worksheet by exact name, falling back to case-insensitive search."""
    if name in wb.sheetnames:
        return wb[name]
    low = name.lower()
    for sn in wb.sheetnames:
        if sn.lower() == low:
            return wb[sn]
    return None


def _fill_appendix_c(wb, d):
    """Insert photos into Appendix-C sheets.

    Photos are inserted WITHOUT burning circles into the JPEG — instead the
    function returns a list of oval descriptors so the caller can inject
    editable Excel AutoShape ovals via _inject_oval_shapes().

    Returns:
        ovals: list of (sheet_name, from_col, from_row, to_col, to_row, shape_id)
               (0-indexed, matching openpyxl/Excel drawing anchor convention)
    """
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor

    photos      = d.get('photos', [])
    categories  = d.get('photo_categories', [])
    titles      = d.get('photo_titles', [])
    coords_list = list(d.get('photo_coords', [])) + [None] * len(photos)

    # 'damaged' is the canonical category; 'damage' (singular) is never used
    sub_photos   = [(p, t, c) for p, cat, t, c in
                    zip(photos, categories, titles + [''] * len(photos), coords_list)
                    if cat == 'general']
    super_photos = [(p, t, c) for p, cat, t, c in
                    zip(photos, categories, titles + [''] * len(photos), coords_list)
                    if cat == 'damaged']

    CAPTION_FILL   = PatternFill(patternType='solid', fgColor='FCE4D6')
    CAPTION_FONT   = Font(name='Times New Roman', size=11, bold=True)
    CAPTION_ALIGN  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    _thin          = Side(style='thin', color='000000')
    CAPTION_BORDER = Border(top=_thin, left=_thin, right=_thin, bottom=_thin)

    ovals     = []
    shape_ctr = [100]   # mutable counter shared across both sheet inserts

    def _insert_photos(ws, photo_list):
        if not photo_list or ws is None:
            return
        ws._images.clear()
        # Clear stale caption merges (keep the header B2:H2)
        for rng in list(ws.merged_cells.ranges):
            if str(rng) != 'B2:H2':
                try:
                    ws.unmerge_cells(str(rng))
                except Exception:
                    pass

        row = 3
        for path, title, coords in photo_list:
            if not path or not os.path.exists(path):
                continue
            try:
                from PIL import Image as PILImage
                with PILImage.open(path) as img:
                    img.load()
                    if img.mode in ('RGBA', 'P', 'LA'):
                        img = img.convert('RGB')
                    # Do NOT burn circle into image — editable shapes injected later
                    w, h  = img.size
                    scale = min(1.0, 1200 / w, 900 / h)
                    new_w = int(w * scale)
                    new_h = int(h * scale)
                    buf   = io.BytesIO()
                    img.resize((new_w, new_h), PILImage.LANCZOS).save(
                        buf, format='JPEG', quality=90)
                buf.seek(0)

                # Photo anchor: rows row-1 → row+19, cols 0 → 8 (0-indexed)
                photo_from_row = row - 1
                photo_to_row   = row + 19
                photo_from_col = 0
                photo_to_col   = 8

                xl_img        = XLImage(buf)
                xl_img.width  = new_w
                xl_img.height = new_h
                anchor        = TwoCellAnchor()
                anchor._from  = AnchorMarker(col=photo_from_col, row=photo_from_row,
                                              colOff=0, rowOff=0)
                anchor.to     = AnchorMarker(col=photo_to_col,   row=photo_to_row,
                                              colOff=0, rowOff=0)
                anchor.editAs = 'oneCell'
                xl_img.anchor = anchor
                ws.add_image(xl_img)

                # If defect coords available, schedule an editable oval shape
                if coords and not _has_red_markers(path):
                    x_pct, y_pct = coords
                    span_cols = photo_to_col - photo_from_col   # 8
                    span_rows = photo_to_row - photo_from_row   # 20
                    # Oval occupies ~14% of photo width/height centred on defect
                    r_cols = max(1, int(span_cols * 0.07))
                    r_rows = max(1, int(span_rows * 0.07))
                    oval_fc = photo_from_col + max(0, int(x_pct * span_cols) - r_cols)
                    oval_tc = photo_from_col + min(span_cols, int(x_pct * span_cols) + r_cols)
                    oval_fr = photo_from_row + max(0, int(y_pct * span_rows) - r_rows)
                    oval_tr = photo_from_row + min(span_rows, int(y_pct * span_rows) + r_rows)
                    shape_ctr[0] += 1
                    ovals.append((ws.title, oval_fc, oval_fr, oval_tc, oval_tr, shape_ctr[0]))

                cap_row = photo_to_row + 2   # two blank rows below photo block
                try:
                    ws.merge_cells(start_row=cap_row, start_column=1,
                                   end_row=cap_row, end_column=8)
                except Exception:
                    pass
                cap_cell           = ws.cell(row=cap_row, column=1, value=title or path)
                cap_cell.fill      = CAPTION_FILL
                cap_cell.font      = CAPTION_FONT
                cap_cell.alignment = CAPTION_ALIGN
                cap_cell.border    = CAPTION_BORDER
                ws.row_dimensions[cap_row].height = 28

                row = cap_row + 3   # gap before next photo

            except Exception as e:
                print(f"EXCEL PHOTO INSERT FAILED {path}: {e}")

    ws_sub   = _find_sheet_rb(wb, 'Appendix__C')
    ws_super = _find_sheet_rb(wb, 'Appendix__C (2)')

    if ws_sub is None:
        print("WARNING: 'Appendix__C' sheet not found — sub-structure photos skipped")
    if ws_super is None:
        print("WARNING: 'Appendix__C (2)' sheet not found — super-structure photos skipped")

    _insert_photos(ws_sub,   sub_photos)
    _insert_photos(ws_super, super_photos)
    return ovals


# ─────────────────────────────────────────────────────────────
#  Editable oval shape injection (post-processing)
# ─────────────────────────────────────────────────────────────

def _inject_oval_shapes(xlsx_path: str, ovals: list):
    """Inject editable red oval AutoShapes into an already-saved xlsx file.

    ovals: list of (sheet_name, from_col, from_row, to_col, to_row, shape_id)
           — from_col/from_row/to_col/to_row are 0-indexed Excel drawing anchors.

    The function:
      1. Opens the xlsx as a ZIP archive.
      2. Resolves each sheet name → its drawing XML file via workbook rels.
      3. Appends <xdr:twoCellAnchor> oval elements to each drawing.
      4. Writes a new xlsx (atomic replace).
    """
    if not ovals:
        return

    import zipfile
    try:
        from lxml import etree
    except ImportError:
        print("lxml not installed — editable ovals skipped (pip install lxml)")
        return

    XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
    A   = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    R   = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    REL = 'http://schemas.openxmlformats.org/package/2006/relationships'

    def _oval_xml(sid, fc, fr, tc, tr):
        return (
            f'<xdr:twoCellAnchor xmlns:xdr="{XDR}" xmlns:a="{A}" editAs="oneCell">'
            f'<xdr:from><xdr:col>{fc}</xdr:col><xdr:colOff>0</xdr:colOff>'
            f'<xdr:row>{fr}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>'
            f'<xdr:to><xdr:col>{tc}</xdr:col><xdr:colOff>0</xdr:colOff>'
            f'<xdr:row>{tr}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>'
            f'<xdr:sp macro="" textlink="">'
            f'<xdr:nvSpPr>'
            f'<xdr:cNvPr id="{sid}" name="DefectCircle{sid}"/>'
            f'<xdr:cNvSpPr><a:spLocks noGrp="1"/></xdr:cNvSpPr>'
            f'</xdr:nvSpPr>'
            f'<xdr:spPr>'
            f'<a:prstGeom prst="ellipse"><a:avLst/></a:prstGeom>'
            f'<a:noFill/>'
            f'<a:ln w="25400"><a:solidFill><a:srgbClr val="FF0000"/></a:solidFill></a:ln>'
            f'</xdr:spPr>'
            f'<xdr:txBody><a:bodyPr rtlCol="0" anchor="ctr"/>'
            f'<a:lstStyle/><a:p/></xdr:txBody>'
            f'</xdr:sp><xdr:clientData/></xdr:twoCellAnchor>'
        )

    # --- Load all files from the xlsx ZIP ---
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        files = {n: z.read(n) for n in z.namelist()}

    # --- Resolve sheet name → drawing file ---
    def _parse(data):
        return etree.fromstring(data)

    wb_ns  = {'wb': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    rel_ns = {'rel': REL}

    wb_xml   = _parse(files['xl/workbook.xml'])
    wb_rels  = _parse(files.get('xl/_rels/workbook.xml.rels', b'<Relationships/>'))

    rid_to_target = {rel.get('Id'): rel.get('Target')
                     for rel in wb_rels.findall('.//rel:Relationship', rel_ns)}

    sheet_to_drawing = {}   # sheet_name → 'xl/drawings/drawingN.xml'
    for sh in wb_xml.findall('.//wb:sheet', wb_ns):
        sname = sh.get('name', '')
        rid   = sh.get(f'{{{R}}}id') or sh.get('r:id', '')
        ws_target = rid_to_target.get(rid, '')
        # ws_target may be 'worksheets/sheet3.xml' (relative to xl/)
        if ws_target.startswith('/xl/'):
            ws_path = ws_target[1:]          # strip leading /
        elif ws_target.startswith('xl/'):
            ws_path = ws_target
        else:
            ws_path = f'xl/{ws_target}'

        ws_fname    = ws_path.split('/')[-1]
        ws_rels_key = f'xl/worksheets/_rels/{ws_fname}.rels'
        if ws_rels_key not in files:
            continue

        ws_rels = _parse(files[ws_rels_key])
        for rel in ws_rels.findall('.//rel:Relationship', rel_ns):
            if 'drawing' in rel.get('Type', '').lower():
                drw = rel.get('Target', '')
                # Target is relative to worksheets/ so '../drawings/drawingN.xml'
                drw_fname = drw.split('/')[-1]
                sheet_to_drawing[sname] = f'xl/drawings/{drw_fname}'
                break

    # --- Group ovals by sheet ---
    ovals_by_sheet = {}
    for sname, fc, fr, tc, tr, sid in ovals:
        ovals_by_sheet.setdefault(sname, []).append((fc, fr, tc, tr, sid))

    # --- Inject ovals into drawing XMLs ---
    modified = False
    for sname, oval_list in ovals_by_sheet.items():
        drw_key = sheet_to_drawing.get(sname)
        if not drw_key or drw_key not in files:
            print(f"OVAL INJECT: drawing not found for sheet '{sname}', skipping")
            continue

        drw_root = _parse(files[drw_key])
        for fc, fr, tc, tr, sid in oval_list:
            oval_el = etree.fromstring(_oval_xml(sid, fc, fr, tc, tr))
            drw_root.append(oval_el)
            print(f"OVAL INJECT: shape {sid} → sheet '{sname}' "
                  f"cols {fc}-{tc}, rows {fr}-{tr}")

        files[drw_key] = etree.tostring(drw_root, xml_declaration=True,
                                         encoding='UTF-8', standalone=True)
        modified = True

    if not modified:
        return

    # --- Write back ---
    tmp = xlsx_path + '._tmp'
    with zipfile.ZipFile(tmp, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)
    os.replace(tmp, xlsx_path)
    print(f"OVAL INJECT: saved {len(ovals)} oval(s) into {xlsx_path}")


def _fill_appendix_c_captions(wb, d):
    """Write the photo caption index to Appendix-c sheet."""
    from openpyxl.cell.cell import MergedCell
    ws     = wb['Appendix-c']
    photos = d.get('photos', [])
    titles = d.get('photo_titles', [])
    cats   = d.get('photo_categories', [])

    # Clear existing content below row 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None

    row = 2
    _safe_write(ws, row, 2, 'SUB STRUCTURE')
    row += 1
    sub_n = 1
    for path, title, cat in zip(photos, titles, cats):
        if cat == 'general':
            _safe_write(ws, row, 2, f'{sub_n}. {title}')
            sub_n += 1
            row += 1

    row += 1
    _safe_write(ws, row, 2, 'SUPER STRUCTURE')
    row += 1
    sup_n = 1
    for path, title, cat in zip(photos, titles, cats):
        if cat == 'damaged':
            _safe_write(ws, row, 2, f'{sup_n}. {title}')
            sup_n += 1
            row += 1


# ─────────────────────────────────────────────────────────────
#  Entry point
# ─────────────────────────────────────────────────────────────

def build_excel(report_json: dict) -> str:
    """Fill CASAD Excel template with report_json and return saved file path."""
    if _TEMPLATE_BYTES:
        wb = openpyxl.load_workbook(io.BytesIO(_TEMPLATE_BYTES))
    else:
        wb = openpyxl.load_workbook(EXCEL_TEMPLATE_PATH)
    try:
        _fill_title_page(wb, report_json)
        _fill_appendix_a(wb, report_json)
        _fill_appendix_b(wb, report_json)
        _fill_defect_tables(wb, report_json)
        _fill_appendix_c_captions(wb, report_json)
        ovals = _fill_appendix_c(wb, report_json)   # returns list of oval descriptors

        name     = re.sub(r'[^\w\-]', '_',
                          report_json.get('bridge_title') or report_json.get('river_name', 'bridge'))
        road     = re.sub(r'[^\w\-]', '_', report_json.get('road_name', 'road'))
        date_str = report_json.get('date_of_survey', 'report').replace('/', '-')
        out_path = os.path.join(OUTPUT_DIR, f'CASAD_{name}_{road}_{date_str}.xlsx')
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        wb.save(out_path)
        print(f"EXCEL REPORT SAVED: {out_path}")
    finally:
        wb.close()

    # Inject editable oval shapes (post-processing — must happen after wb.save)
    if ovals:
        _inject_oval_shapes(out_path, ovals)

    return out_path
