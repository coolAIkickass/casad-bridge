"""
Generate CASAD Bridge Inspection Bot — Cost Analysis Excel sheet.
Run: python3 generate_cost_sheet.py
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
import datetime

# ── Colour palette ────────────────────────────────────────────────────────────
NAV   = "1F3864"   # dark navy   (headers)
BLUE  = "2E75B6"   # mid blue    (sub-headers)
LBLUE = "D6E4F0"   # light blue  (alt rows / section fills)
VBLUE = "EBF5FB"   # very light  (zebra rows)
GREEN = "E2EFDA"   # light green (totals)
DGRN  = "375623"   # dark green  (total text)
YELL  = "FFF2CC"   # yellow      (notes / assumptions)
WHITE = "FFFFFF"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=WHITE, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def _border(style="thin"):
    s = Side(style=style, color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _money(ws, cell_ref, value):
    ws[cell_ref] = value
    ws[cell_ref].number_format = '"$"#,##0.00'

def _pct(ws, cell_ref, value):
    ws[cell_ref] = value / 100
    ws[cell_ref].number_format = '0%'


def style_header_row(ws, row, col_start, col_end, text, bg=NAV, font_size=12):
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row,   end_column=col_end
    )
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.fill      = _fill(bg)
    cell.font      = _font(bold=True, color=WHITE, size=font_size)
    cell.alignment = _center(wrap=True)
    cell.border    = _border()


def style_subheader(ws, row, cols, texts, bg=BLUE):
    for col, text in zip(cols, texts):
        c = ws.cell(row=row, column=col, value=text)
        c.fill      = _fill(bg)
        c.font      = _font(bold=True, color=WHITE, size=10)
        c.alignment = _center(wrap=True)
        c.border    = _border()


def style_data_row(ws, row, cols, values, bg=WHITE, bold=False,
                   number_cols=None, fmt=None):
    number_cols = number_cols or []
    for col, val in zip(cols, values):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = _fill(bg)
        c.font      = _font(bold=bold, color="1F1F1F" if bg != NAV else WHITE,
                            size=10)
        c.alignment = _center(wrap=True) if col > 1 else _left()
        c.border    = _border()
        if col in number_cols and isinstance(val, (int, float)):
            c.number_format = '"$"#,##0.00'
        if fmt and col in fmt:
            c.number_format = fmt[col]


wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Executive Summary
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 38
for col in ["B", "C", "D", "E"]:
    ws1.column_dimensions[col].width = 18
ws1.row_dimensions[1].height = 14
ws1.row_dimensions[2].height = 42
ws1.row_dimensions[3].height = 22

# Title
style_header_row(ws1, 2, 1, 5,
    "CASAD CONSULTANTS PVT. LTD.\nBridge Inspection Bot — Cost Analysis",
    bg=NAV, font_size=14)

# Prepared by / date
ws1.merge_cells("A3:E3")
c = ws1["A3"]
c.value     = f"Prepared for: Founder Sign-off   |   Date: {datetime.date.today().strftime('%d %B %Y')}   |   Prepared by: Technology Team"
c.fill      = _fill(LBLUE)
c.font      = _font(bold=False, color=NAV, size=10, italic=True)
c.alignment = _center()
c.border    = _border()

# ── Section A: Usage Assumptions ─────────────────────────────────────────────
row = 5
style_header_row(ws1, row, 1, 5, "A.  USAGE ASSUMPTIONS", bg=BLUE, font_size=11)
row += 1
style_subheader(ws1, row, [1,2,3,4,5],
    ["Parameter", "Off-Season", "Peak Season", "Max Burst", "Unit"])
row += 1

assumptions = [
    ("Concurrent users / day",        10,   10,   10,  "users"),
    ("Reports / month",               25,   80,   80,  "reports"),
    ("Avg photos / report (typical)", 25,   25,   25,  "photos"),
    ("Max photos / report (heavy)",   60,   60,   60,  "photos"),
    ("Photos with captions (defect)", "~60%","~60%","~60%", "of total photos"),
    ("Audio voice notes / report",    3,    5,    8,   "notes"),
    ("Season duration",               10,   2,    "--", "months/year"),
]
for i, (param, off, peak, burst, unit) in enumerate(assumptions):
    bg = VBLUE if i % 2 == 0 else WHITE
    style_data_row(ws1, row, [1,2,3,4,5],
                   [param, off, peak, burst, unit], bg=bg)
    row += 1

# ── Section B: Per-Report Cost Breakdown ─────────────────────────────────────
row += 1
style_header_row(ws1, row, 1, 5, "B.  PER-REPORT COST BREAKDOWN", bg=BLUE, font_size=11)
row += 1
style_subheader(ws1, row, [1,2,3,4,5],
    ["Service / Component", "Typical (25 photos)", "Heavy (60 photos)",
     "Free Tier?", "Notes"])
row += 1

per_report = [
    ("Claude Sonnet — Field note parsing\n(ai_parse.py, once per report)",
     0.05, 0.08, "No", "~4,500 tok in / 2,500 tok out (typical)\n~6,000 / 4,000 tok (heavy)"),
    ("Claude Haiku — Defect circle detection\n(mark_image.py, per photo w/ caption)",
     0.045, 0.11, "No", "~1,700 tok/photo; 60% photos have captions\n15 calls (typical) / 36 calls (heavy)"),
    ("Groq — Audio transcription\n(whisper-large-v3)",
     0.00, 0.00, "Yes", "Free tier covers thousands of mins/month;\nno cost at this volume"),
    ("WhatsApp Cloud API\n(messages + document delivery)",
     0.00, 0.00, "Yes*", "First 1,000 conversations/month free.\n80 reports = well under limit."),
    ("Render — Server hosting\n(allocated per report)",
     0.00, 0.00, "Fixed", "Fixed $7/month; allocated separately below"),
]
for i, (svc, typ, heavy, free, note) in enumerate(per_report):
    bg = VBLUE if i % 2 == 0 else WHITE
    style_data_row(ws1, row, [1,2,3,4,5],
                   [svc, typ, heavy, free, note],
                   bg=bg, number_cols=[2,3])
    row += 1

# Total row
style_data_row(ws1, row, [1,2,3,4,5],
               ["TOTAL (API cost per report)", 0.095, 0.19, "", "Excl. fixed server cost"],
               bg=GREEN, bold=True, number_cols=[2,3])
ws1.cell(row=row, column=1).font = _font(bold=True, color=DGRN, size=10)
ws1.cell(row=row, column=2).font = _font(bold=True, color=DGRN, size=10)
ws1.cell(row=row, column=3).font = _font(bold=True, color=DGRN, size=10)

# ── Section C: Monthly Cost Summary ──────────────────────────────────────────
row += 2
style_header_row(ws1, row, 1, 5, "C.  MONTHLY COST SUMMARY", bg=BLUE, font_size=11)
row += 1
style_subheader(ws1, row, [1,2,3,4,5],
    ["Cost Component", "Off-Season\n(~25 reports)", "Peak Season\n(~80 reports)",
     "Worst Case\n(80 × 60 photos)", "Frequency"])
row += 1

monthly = [
    ("Claude API (variable)",   2.38,  7.60,  15.20, "Monthly"),
    ("Groq (audio)",            0.00,  0.00,   0.00, "Monthly — Free"),
    ("WhatsApp Cloud API",      0.00,  0.00,   0.00, "Monthly — Free ≤1000 conv"),
    ("Render (server hosting)", 7.00,  7.00,   7.00, "Monthly — Fixed"),
    ("TOTAL / month",           9.38, 14.60,  22.20, ""),
]
for i, (item, off, peak, worst, freq) in enumerate(monthly):
    is_total = item.startswith("TOTAL")
    bg = GREEN if is_total else (VBLUE if i % 2 == 0 else WHITE)
    style_data_row(ws1, row, [1,2,3,4,5],
                   [item, off, peak, worst, freq],
                   bg=bg, bold=is_total, number_cols=[2,3,4])
    if is_total:
        for col in [1,2,3,4]:
            ws1.cell(row=row, column=col).font = _font(bold=True, color=DGRN, size=10)
    row += 1

# ── Section D: Annual Projection ─────────────────────────────────────────────
row += 1
style_header_row(ws1, row, 1, 5, "D.  ANNUAL COST PROJECTION", bg=BLUE, font_size=11)
row += 1
style_subheader(ws1, row, [1,2,3,4,5],
    ["Period", "Duration", "Monthly Cost", "Sub-Total", "Notes"])
row += 1

annual = [
    ("Off-Season months",  "10 months", 9.38,   93.80, "~25 reports/month"),
    ("Peak Season months", "2 months",  14.60,  29.20, "~80 reports/month"),
    ("ANNUAL TOTAL",       "12 months", "--",  123.00, "~$10.25 / month avg"),
]
for i, (period, dur, monthly_c, sub, note) in enumerate(annual):
    is_total = period == "ANNUAL TOTAL"
    bg = GREEN if is_total else (VBLUE if i % 2 == 0 else WHITE)
    vals = [period, dur, monthly_c if not is_total else "--", sub, note]
    style_data_row(ws1, row, [1,2,3,4,5], vals,
                   bg=bg, bold=is_total,
                   number_cols=[3,4] if not is_total else [4])
    if is_total:
        for col in [1,2,4]:
            ws1.cell(row=row, column=col).font = _font(bold=True, color=DGRN, size=10)
    row += 1

# ── Section E: Key Notes ──────────────────────────────────────────────────────
row += 1
style_header_row(ws1, row, 1, 5, "E.  KEY NOTES & ASSUMPTIONS", bg=BLUE, font_size=11)
row += 1

notes = [
    "1. Claude API pricing used: Sonnet $3/M input, $15/M output; Haiku $1/M input, $5/M output.",
    "2. WhatsApp Cloud API: first 1,000 conversations/month are free. At 80 reports/month the system stays within the free tier.",
    "3. Groq (Whisper) free tier covers thousands of audio minutes per month — no cost expected at this scale.",
    "4. Render Starter plan ($7/month) is a fixed cost regardless of usage volume.",
    "5. 'Heavy report' = 60 photos, all with captions triggering defect-circle AI detection.",
    "6. Costs scale linearly with report volume; the fixed Render cost dominates at low volumes.",
    "7. No setup or one-time costs assumed; all figures are recurring operational costs.",
]
for i, note in enumerate(notes):
    ws1.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=5
    )
    c = ws1.cell(row=row, column=1, value=note)
    c.fill      = _fill(YELL if i % 2 == 0 else "FFFAE5")
    c.font      = _font(bold=False, color="3D3D00", size=9)
    c.alignment = _left()
    c.border    = _border()
    ws1.row_dimensions[row].height = 20
    row += 1


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Monthly Scenario Detail
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Monthly Scenarios")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 32
for col in ["B","C","D","E","F"]:
    ws2.column_dimensions[col].width = 16

style_header_row(ws2, 1, 1, 6,
    "Monthly Cost Scenarios — Detailed Breakdown", bg=NAV, font_size=13)
style_header_row(ws2, 2, 1, 6,
    "Assumptions: Sonnet $3/$15 per M tokens | Haiku $1/$5 per M tokens | 60% photos have captions",
    bg=LBLUE, font_size=9)

ws2.cell(row=2, column=1).font = _font(bold=False, color=NAV, size=9, italic=True)

row = 4
style_subheader(ws2, row, [1,2,3,4,5,6],
    ["Line Item", "Off-Season\n(20 reports)", "Off-Season\n(30 reports)",
     "Peak\n(60 reports)", "Peak\n(80 reports)", "Worst Case\n(80 × 60 photos)"])
row += 1

detail_rows = [
    # (label, [20, 30, 60, 80, 80-heavy])
    ("Reports / month",             [20,    30,    60,    80,    80   ]),
    ("Avg photos / report",         [25,    25,    25,    25,    60   ]),
    ("Total photos",                [500,   750,   1500,  2000,  4800 ]),
    ("Photos with captions (60%)",  [300,   450,   900,   1200,  2880 ]),
    ("── Claude Sonnet cost",       [1.00,  1.50,  3.00,  4.00,  6.40 ]),
    ("── Claude Haiku cost",        [0.90,  1.35,  2.70,  3.60,  8.64 ]),
    ("── Groq (audio)",             [0.00,  0.00,  0.00,  0.00,  0.00 ]),
    ("── WhatsApp API",             [0.00,  0.00,  0.00,  0.00,  0.00 ]),
    ("── Render (server)",          [7.00,  7.00,  7.00,  7.00,  7.00 ]),
    ("TOTAL / month",               [8.90, 9.85,  12.70, 14.60, 22.04]),
    ("Cost per report",             [0.445, 0.328, 0.212, 0.183, 0.275]),
]

money_rows = {"── Claude Sonnet cost", "── Claude Haiku cost", "── Groq (audio)",
              "── WhatsApp API", "── Render (server)", "TOTAL / month", "Cost per report"}

for i, (label, vals) in enumerate(detail_rows):
    is_total  = label == "TOTAL / month"
    is_cost_pr= label == "Cost per report"
    is_section= label.startswith("──")
    bg = GREEN if is_total else (LBLUE if is_cost_pr else (VBLUE if i % 2 == 0 else WHITE))

    row_vals = [label] + vals
    style_data_row(ws2, row, list(range(1, 7)), row_vals,
                   bg=bg, bold=is_total or is_cost_pr,
                   number_cols=list(range(2, 7)) if label in money_rows else [])
    if is_total or is_cost_pr:
        for col in range(1, 7):
            ws2.cell(row=row, column=col).font = _font(bold=True, color=DGRN, size=10)
    row += 1

# ── Chart ─────────────────────────────────────────────────────────────────────
chart = BarChart()
chart.type        = "col"
chart.title       = "Monthly Total Cost by Scenario"
chart.y_axis.title = "Cost (USD)"
chart.x_axis.title = "Scenario"
chart.style       = 10
chart.width       = 20
chart.height      = 12

# Total row is at row 14 (row 4 header + 10 data rows = row 14)
total_row = 4 + len(detail_rows)  # header row + all data rows
data = Reference(ws2, min_col=2, max_col=6, min_row=total_row, max_row=total_row)
cats = Reference(ws2, min_col=2, max_col=6, min_row=4, max_row=4)
chart.add_data(data, titles_from_data=False)
chart.set_categories(cats)

from openpyxl.chart.series import SeriesLabel
chart.series[0].title = SeriesLabel(v="Total Monthly Cost ($)")

ws2.add_chart(chart, f"A{total_row + 3}")


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — Annual Projection
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Annual Projection")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 22
for col in ["B","C","D","E","F","G","H","I","J","K","L","M","N"]:
    ws3.column_dimensions[col].width = 11

style_header_row(ws3, 1, 1, 14, "Annual Cost Projection — Month by Month", bg=NAV, font_size=13)

months = ["Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar","Annual"]
# Peak months assumed: Jun, Jul, Aug (monsoon inspection season)
is_peak = [False,False,True,True,True,False,False,False,False,False,False,False,False]
reports  = [25,25,80,80,80,25,25,25,25,25,25,25,0]  # 0 = calculated

row = 3
style_subheader(ws3, row, list(range(1,15)), [""] + months)
row += 1

# Build rows
claude_costs  = [7.60 if pk else 2.38 for pk in is_peak] + [0]
render_costs  = [7.00]*12 + [0]
total_monthly = [c + r for c, r in zip(claude_costs, render_costs)] + [0]

annual_claude = sum(claude_costs)
annual_render = sum(render_costs)
annual_total  = annual_claude + annual_render

claude_costs[-1]  = annual_claude
render_costs[-1]  = annual_render
total_monthly[-1] = annual_total

report_row  = ["Reports/month"] + [str(r) for r in reports[:-1]] + [str(sum(reports[:-1]))]
season_row  = ["Season"] + ["PEAK" if pk else "Normal" for pk in is_peak] + [""]
claude_row  = ["Claude API ($)"] + claude_costs
render_row  = ["Render ($)"] + render_costs
total_row_v = ["TOTAL ($)"] + total_monthly

table_data = [
    (report_row,  WHITE,  False),
    (season_row,  VBLUE,  False),
    (claude_row,  WHITE,  True),
    (render_row,  VBLUE,  True),
    (total_row_v, GREEN,  True),
]

for tdata, bg, is_money in table_data:
    is_total = tdata[0] == "TOTAL ($)"
    for col_idx, val in enumerate(tdata, 1):
        c = ws3.cell(row=row, column=col_idx, value=val)
        # Highlight peak columns
        col_is_peak = col_idx >= 4 and col_idx <= 6  # Jul, Aug, Sep (cols 4-6 = peak)
        cell_bg = "FFE699" if (col_is_peak and not is_total) else bg
        if is_total and col_is_peak:
            cell_bg = "C6EFCE"
        c.fill      = _fill(cell_bg)
        c.font      = Font(bold=is_total, color=DGRN if is_total else "1F1F1F",
                           size=9, name="Calibri")
        c.alignment = _center()
        c.border    = _border()
        if is_money and col_idx > 1 and isinstance(val, float):
            c.number_format = '"$"#,##0.00'
    row += 1

# Highlight header for peak months
for col in [4, 5, 6]:  # Jun, Jul, Aug
    c = ws3.cell(row=3, column=col)
    c.fill = _fill("ED7D31")
    c.font = _font(bold=True, color=WHITE, size=10)

# Note row
row += 1
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
c = ws3.cell(row=row, column=1,
    value="* Peak season assumed Jun–Aug (monsoon inspection period). Yellow = peak month columns.")
c.fill = _fill(YELL)
c.font = Font(italic=True, color="3D3D00", size=9, name="Calibri")
c.alignment = _left()
c.border = _border()


# ══════════════════════════════════════════════════════════════════════════════
# Save
# ══════════════════════════════════════════════════════════════════════════════
out = "/Users/bansal.umang89/Desktop/CASAD_Bridge_Bot_Cost_Analysis.xlsx"
wb.save(out)
print(f"Saved: {out}")
